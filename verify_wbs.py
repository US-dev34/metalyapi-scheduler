import openpyxl
from openpyxl.utils import get_column_letter
import os

FILE_PATH = os.path.join("U:" + os.sep, "Antigravity", "Antigravity", "Construction Schedule", "DATA", "WBS_E2NS_D1D2_Restructured.xlsx")
SHEET_NAME = "E2NS_WBS"
sep = "=" * 120

print(sep)
print("WBS FILE VERIFICATION REPORT")
print(sep)
print("File:", FILE_PATH)
print("Exists:", os.path.exists(FILE_PATH))
print("Size:", f"{os.path.getsize(FILE_PATH):,}", "bytes")
print()

wb = openpyxl.load_workbook(FILE_PATH)
print("Sheets:", wb.sheetnames)
ws = wb[SHEET_NAME]
print("Dims:", ws.dimensions, "Max row:", ws.max_row, "Max col:", ws.max_column)
print()

print(sep)
print("1. HEADER ROW (Row 4) - Columns A through V")
print(sep)

headers = {}
for col in range(1, 23):
    cell = ws.cell(row=4, column=col)
    cl = get_column_letter(col)
    headers[cl] = cell.value
    print(f"  Column {cl:>2} (col {col:>2}): {cell.value}")

print()
print("New columns J-V:")
for c in range(10, 23):
    cl = get_column_letter(c)
    val = ws.cell(row=4, column=c).value
    st = "OK" if val is not None else "EMPTY"
    print(f"  {cl}: {val}  [{st}]")

print()
print(sep)
print("2. FIRST 15 DATA ROWS (rows 5-19)")
print(sep)

hl = "Row |"
for col in range(1, 23):
    cl = get_column_letter(col)
    hn = str(headers.get(cl, ""))[:12]
    hl += f" {cl}:{hn:>12} |"
print(hl)
print("-" * len(hl))

for row in range(5, 20):
    rd = f"{row:>3} |"
    for col in range(1, 23):
        val = ws.cell(row=row, column=col).value
        vs = str(val)[:12] if val is not None else ""
        rd += f" {vs:>14} |"
    print(rd)

print()
print(sep)
print("3. YELLOW FILL (FFFF00) COUNT")
print(sep)

yc = 0
ys = []
ybc = {}

for col in range(1, ws.max_column + 1):
    cl = get_column_letter(col)
    cy = 0
    for row in range(5, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if "FFFF00" in rgb:
                cy += 1
                yc += 1
                if len(ys) < 10:
                    ys.append((row, cl, cell.value))
    if cy > 0:
        ybc[cl] = cy

print(f"Total yellow (FFFF00) cells: {yc}")
if ybc:
    print("By column:")
    for cl, cnt in ybc.items():
        print(f"  Column {cl}: {cnt}")
else:
    print("No exact FFFF00. Scanning fills in J-V...")
    fc = {}
    for col in range(10, 23):
        for row in range(5, min(ws.max_row + 1, 50)):
            cell = ws.cell(row=row, column=col)
            fill = cell.fill
            if fill.patternType and fill.patternType != "none":
                fg = str(fill.fgColor.rgb) if fill.fgColor and fill.fgColor.rgb else "no-fg"
                bg = str(fill.bgColor.rgb) if fill.bgColor and fill.bgColor.rgb else "no-bg"
                key = f"pat={fill.patternType} fg={fg} bg={bg}"
                if key not in fc: fc[key] = []
                fc[key].append((row, get_column_letter(col)))
    if fc:
        for color, cells in fc.items():
            print(f"  {color}: {len(cells)} cells (e.g. {cells[:5]})")
    else:
        print("  No fills found.")

if ys:
    print("Sample yellow cells:")
    for row, col, val in ys:
        print(f"  Row {row}, Col {col}: value = {val}")

print()
print(sep)
print("4. QTY CELLS = 1 COUNT")
print(sep)

qty_col = None
for col in range(1, 23):
    hdr = ws.cell(row=4, column=col).value
    if hdr and str(hdr).strip().upper() == "QTY":
        qty_col = col
        break
if qty_col is None:
    for col in range(1, 23):
        hdr = ws.cell(row=4, column=col).value
        if hdr and "qty" in str(hdr).lower():
            qty_col = col
            break

if qty_col:
    ql = get_column_letter(qty_col)
    print(f"QTY column: {ql} (col {qty_col})")
    hv = ws.cell(row=4, column=qty_col).value
    print("Header:", repr(hv))
    q1 = qo = qe = 0
    qv = {}
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=qty_col).value
        if val is None or str(val).strip() == "":
            qe += 1
        elif val == 1 or str(val).strip() == "1":
            q1 += 1
        else:
            qo += 1
            v = str(val)
            qv[v] = qv.get(v, 0) + 1
    print(f"  QTY=1: {q1}")
    print(f"  QTY=other: {qo}")
    print(f"  QTY=empty: {qe}")
    if qv:
        print("  Other values:")
        for v, c in sorted(qv.items(), key=lambda x: -x[1]):
            print(f"    {v}: {c}")
else:
    print("QTY column not found!")

print()
print(sep)
print("5. ROWS WITH MANPOWER / DURATION / TOTAL MANDAY DATA")
print(sep)

mc = dc = tc = None
for col in range(1, 23):
    hdr = ws.cell(row=4, column=col).value
    if hdr:
        h = str(hdr).lower()
        if "manpower" in h: mc = col
        if "duration" in h and "total" not in h: dc = col
        if "total" in h and "manday" in h: tc = col

for label, c in [("Manpower", mc), ("Duration", dc), ("Total Manday", tc)]:
    if c:
        print(f"{label}: {get_column_letter(c)} (col {c})")
    else:
        print(f"{label}: NOT FOUND")

rm = rd = rt = 0
sr = []
for row in range(5, ws.max_row + 1):
    mp = ws.cell(row=row, column=mc).value if mc else None
    du = ws.cell(row=row, column=dc).value if dc else None
    tm = ws.cell(row=row, column=tc).value if tc else None
    hm = mp is not None and str(mp).strip() != ""
    hd = du is not None and str(du).strip() != ""
    ht = tm is not None and str(tm).strip() != ""
    if hm: rm += 1
    if hd: rd += 1
    if ht: rt += 1
    if (hm or hd or ht) and len(sr) < 10:
        sr.append((row, mp, du, tm))

print(f"Rows with Manpower: {rm}")
print(f"Rows with Duration: {rd}")
print(f"Rows with Total Manday: {rt}")

if sr:
    print("Sample rows:")
    for row, mp, du, tm in sr:
        print(f"  Row {row}: Manpower={mp}, Duration={du}, TotalManday={tm}")

print()
print(sep)
print("VERIFICATION COMPLETE")
print(sep)
wb.close()
