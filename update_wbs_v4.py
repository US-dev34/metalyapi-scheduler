import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

src = r"U:\Antigravity\Antigravity\Construction Schedule\DATA\WBS_E2NS_D1D2_Restructured.xlsx"
tp_path = r"U:\Antigravity\Antigravity\Construction Schedule\DATA\Terminplan\260113_DRAFT_E2NS Terminplan.xlsx"
backup = r"U:\Antigravity\Antigravity\Construction Schedule\DATA\WBS_E2NS_D1D2_Restructured_BACKUP_v4.xlsx"

shutil.copy2(src, backup)
print("Backup created:", backup)

# ============================================================
# STEP 1: Read Terminplan data
# Cols: E=Pos, F=Desc, J=QTY, K=Done, L=Rem, M=Manpower, N=Duration, O=TotalManday
# ============================================================
tp_wb = openpyxl.load_workbook(tp_path, data_only=True)
tp_ws = tp_wb['E2NS Terminplan']

tp_data = {}
for row in tp_ws.iter_rows(min_row=12, max_row=tp_ws.max_row, max_col=20):
    pos = row[4].value   # col E (index 4)
    desc = row[5].value  # col F
    qty = row[9].value   # col J
    done = row[10].value # col K
    rem = row[11].value  # col L
    manpower = row[12].value  # col M
    duration = row[13].value  # col N
    manday = row[14].value    # col O

    if pos is not None:
        tp_data[str(pos)] = {
            'pos': str(pos), 'desc': str(desc or '').strip(),
            'qty': qty, 'done': done, 'rem': rem,
            'manpower': manpower, 'duration': duration, 'manday': manday
        }

print("Terminplan: %d positions loaded" % len(tp_data))

# ============================================================
# STEP 2: MAPPING - same as v3
# ============================================================
MAPPING = {
    # === E2N (1.0.0) ===
    '1.1.0': ('1.1.3.1', 'EXT'), '1.1.1': ('1.1.3.1', 'EXT'), '1.1.2': ('1.1.3.1', 'EXT'),
    '1.1.3': ('1.1.3.1', 'EXT'), '1.1.4': ('1.1.3.1', 'EXT'),
    '1.2.0': ('1.1.1.2', 'EXT'), '1.2.1': ('1.1.1.2', 'EXT'), '1.2.2': ('1.1.1.2', 'EXT'),
    '1.2.3': ('1.1.1.2', 'EXT'),
    '1.3.0': ('1.1.2.1', 'EXT'),
    '1.4.0': ('1.1.4', 'EXT'), '1.4.1': ('1.1.4', 'EXT'), '1.4.2': ('1.1.4', 'EXT'),
    '1.5.0': ('1.1.1.1', 'EXT'),
    '1.6.0': ('1.1.2', 'EXT'),
    '1.7.0': ('1.1.7.3', 'EXT'),
    '1.8.0': ('1.1.5', 'EXT'),
    '1.9.0': ('1.1.8.1', 'EXT'),
    '1.10.0': ('1.1.1.4', 'EXT'),
    '1.11.0': ('1.1.10', 'EXT'),
    '1.12.0': ('1.1.6', 'EXT'), '1.12.1': ('1.1.6', 'EXT'), '1.12.2': ('1.1.6', 'EXT'),
    '1.13.0': ('1.1.9', 'EXT'), '1.13.1': ('1.1.9.1', 'EXT'), '1.13.2': ('1.1.9.2', 'EXT'),
    '1.13.3': ('1.1.9.3', 'EXT'),
    '1.14.0': ('1.1.12', 'EXT'), '1.14.1': ('1.1.12', 'EXT'), '1.14.2': ('1.1.12', 'EXT'),
    '1.14.3': ('1.1.12', 'EXT'),
    # === E2S Exterior (2.0.0) ===
    '2.1.0': ('1.2.1.1', 'EXT'), '2.1.1': ('1.2.1.1', 'EXT'), '2.1.2': ('1.2.1.1', 'EXT'),
    '2.1.3': ('1.2.1.1', 'EXT'),
    '2.2.0': ('1.2.4.1', 'EXT'), '2.2.1': ('1.2.4.1', 'EXT'),
    '2.3.0': ('1.2.2.1', 'EXT'), '2.3.1': ('1.2.2.1', 'EXT'), '2.3.2': ('1.2.2.1', 'EXT'),
    '2.3.3': ('1.2.2.1', 'EXT'),
    '2.4.0': ('1.1.8.2', 'EXT'), '2.4.1': ('1.1.8.2', 'EXT'), '2.4.2': ('1.1.8.2', 'EXT'),
    '2.4.3': ('1.1.8.2', 'EXT'), '2.4.3.1': ('1.1.8.2', 'EXT'), '2.4.3.2': ('1.1.8.2', 'EXT'),
    '2.4.3.3': ('1.1.8.2', 'EXT'),
    '2.5.0': ('1.4.1.2', 'EXT'),
    '2.6.0': ('1.2.7', 'EXT'), '2.6.1': ('1.2.7', 'EXT'), '2.6.2': ('1.2.7', 'EXT'),
    '2.6.3': ('1.2.7', 'EXT'),
    '2.7.0': ('1.2.5', 'EXT'),
    '2.8.0': ('1.4.3.1', 'EXT'), '2.8.1': ('1.4.3.1', 'EXT'), '2.8.2': ('1.4.3.1', 'EXT'),
    '2.8.3': ('1.4.3.1', 'EXT'), '2.8.4': ('1.4.3.2', 'EXT'),
    '2.9.0': ('1.4.3.3', 'EXT'), '2.9.1': ('1.4.3.3', 'EXT'), '2.9.2': ('1.4.3.3', 'EXT'),
    '2.9.3': ('1.4.3.3', 'EXT'), '2.9.4': ('1.4.3.4', 'EXT'),
    '2.10.0': ('1.1.11', 'EXT'),
    '2.11.0': ('1.2.2.4', 'EXT'),
    '2.12.0': ('1.2.6', 'EXT'), '2.12.1': ('1.2.6', 'EXT'), '2.12.2': ('1.2.6', 'EXT'),
    '2.12.3': ('1.2.6', 'EXT'),
    '2.13.0': ('1.2.5', 'EXT'), '2.13.1': ('1.2.5', 'EXT'), '2.13.2': ('1.2.5', 'EXT'),
    # === E2S Inenhof (3.0.0) - EXT ===
    '3.1.0': ('1.2.3', 'EXT'), '3.1.1': ('1.2.3.1', 'EXT'), '3.1.2': ('1.2.3.2', 'EXT'),
    '3.1.3': ('1.2.3.3', 'EXT'), '3.1.4': ('1.2.3.4', 'EXT'),
    '3.2.0': ('1.2.4.1', 'EXT'), '3.2.1': ('1.2.4.1', 'EXT'), '3.2.2': ('1.2.4.1', 'EXT'),
    '3.3.0': ('1.2.1.2', 'EXT'), '3.3.1': ('1.2.1.2', 'EXT'), '3.3.2': ('1.2.1.2', 'EXT'),
    '3.3.3': ('1.2.1.2', 'EXT'), '3.3.4': ('1.2.1.2', 'EXT'), '3.3.5': ('1.2.1.2', 'EXT'),
    '3.4.0': ('1.1.7.1', 'EXT'), '3.4.1': ('1.1.7.1', 'EXT'), '3.4.2': ('1.1.7.1', 'EXT'),
    '3.4.3': ('1.1.7.1', 'EXT'), '3.4.4': ('1.1.7.1', 'EXT'), '3.4.5': ('1.1.7.1', 'EXT'),
    '3.4.6': ('1.1.7.1', 'EXT'),
    '3.5.0': ('1.2.8', 'EXT'), '3.5.1': ('1.2.8', 'EXT'), '3.5.2': ('1.2.8', 'EXT'),
    '3.5.3': ('1.2.8', 'EXT'), '3.5.4': ('1.2.8', 'EXT'), '3.5.5': ('1.2.8', 'EXT'),
    '3.5.6': ('1.2.8', 'EXT'), '3.5.7': ('1.2.8', 'EXT'),
    '3.6.0': ('1.2.9', 'EXT'), '3.6.1': ('1.2.9', 'EXT'), '3.6.2': ('1.2.9', 'EXT'),
    '3.6.3': ('1.2.9', 'EXT'), '3.6.4': ('1.2.9', 'EXT'),
    # === E2S Interior (4.0.0) ===
    '4.1.0': ('1.1.2.6', 'INT'), '4.1.1': ('1.1.2.6', 'INT'), '4.1.2': ('1.1.2.6', 'INT'),
    '4.1.3': ('1.1.2.6', 'INT'), '4.1.4': ('1.1.2.6', 'INT'),
    '4.2.0': ('1.2.3', 'INT'), '4.2.1': ('1.2.3.5', 'INT'), '4.2.2': ('1.2.3.6', 'INT'),
    '4.2.3': ('1.2.3.7', 'INT'),
    '4.3.0': ('1.2.4.1', 'INT'), '4.3.1': ('1.2.4.1', 'INT'), '4.3.2': ('1.2.4.1', 'INT'),
    '4.3.3': ('1.2.4.1', 'INT'),
    '4.4.0': ('1.2.1.1', 'INT'), '4.4.1': ('1.2.1.1', 'INT'), '4.4.2': ('1.2.1.1', 'INT'),
    '4.4.3': ('1.2.1.1', 'INT'),
    '4.5.0': ('1.2.1.4', 'INT'), '4.5.1': ('1.2.1.4', 'INT'), '4.5.2': ('1.2.1.4', 'INT'),
    '4.6.0': ('1.2.2.1', 'INT'), '4.6.1': ('1.2.2.1', 'INT'),
    '4.7.0': ('1.2.1.2', 'INT'), '4.7.1': ('1.2.1.2', 'INT'), '4.7.2': ('1.2.1.2', 'INT'),
    '4.7.3': ('1.2.1.2', 'INT'),
    '4.8.0': ('1.2.4.2', 'INT'), '4.8.1': ('1.2.4.2', 'INT'),
    '4.9.0': ('1.2.2.2', 'INT'), '4.9.1': ('1.2.2.2', 'INT'), '4.9.2': ('1.2.2.2', 'INT'),
    '4.9.3': ('1.2.2.2', 'INT'), '4.9.4': ('1.2.2.2', 'INT'),
    '4.10.0': ('1.2.1.2', 'INT'), '4.10.1': ('1.2.1.2', 'INT'), '4.10.2': ('1.2.1.2', 'INT'),
    '4.10.3': ('1.2.1.2', 'INT'), '4.10.4': ('1.2.1.2', 'INT'), '4.10.5': ('1.2.1.2', 'INT'),
    '4.11.0': ('1.2.10', 'INT'),
}

# ============================================================
# STEP 3: Aggregate per WBS + scope
# Now includes: qty, done, rem, manpower, duration, manday
# ============================================================
def num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    return 0

wbs_agg = {}

for tp_pos, (wbs_code, scope) in MAPPING.items():
    if wbs_code is None:
        continue
    tp = tp_data.get(tp_pos)
    if tp is None:
        continue

    if wbs_code not in wbs_agg:
        wbs_agg[wbs_code] = {
            'EXT': {'qty': 0, 'done': 0, 'rem': 0, 'manpower': 0, 'duration': 0, 'manday': 0, 'has_data': False},
            'INT': {'qty': 0, 'done': 0, 'rem': 0, 'manpower': 0, 'duration': 0, 'manday': 0, 'has_data': False}
        }

    q = num(tp['qty'])
    d = num(tp['done'])
    r = num(tp['rem'])
    mp = num(tp['manpower'])
    dur = num(tp['duration'])
    md = num(tp['manday'])

    # If any numeric data exists, mark as has_data
    if q > 0 or d > 0 or r > 0 or mp > 0 or dur > 0 or md > 0:
        a = wbs_agg[wbs_code][scope]
        a['qty'] += q
        a['done'] += d
        a['rem'] += r
        # For manpower: take max (since parallel teams)
        if mp > a['manpower']:
            a['manpower'] = mp
        a['duration'] += dur
        a['manday'] += md
        a['has_data'] = True

print("WBS aggregation: %d codes" % len(wbs_agg))

# ============================================================
# STEP 4: Process WBS file
# ============================================================
wb = openpyxl.load_workbook(src)

# New column layout (after original 9 cols A-I):
# J=Scope, K=QTY_EXT, L=Done_EXT, M=Rem_EXT, N=Manpower_EXT, O=Duration_EXT, P=Manday_EXT,
# Q=QTY_INT, R=Done_INT, S=Rem_INT, T=Manpower_INT, U=Duration_INT, V=Manday_INT
NEW_COLS = 13
TOTAL_COLS = 9 + NEW_COLS  # 22

# Column indices (0-based in our arrays)
COL_SCOPE = 9
COL_QTY_E = 10
COL_DONE_E = 11
COL_REM_E = 12
COL_MP_E = 13
COL_DUR_E = 14
COL_MD_E = 15
COL_QTY_I = 16
COL_DONE_I = 17
COL_REM_I = 18
COL_MP_I = 19
COL_DUR_I = 20
COL_MD_I = 21

HEADERS = {
    COL_SCOPE: 'Scope',
    COL_QTY_E: 'QTY (EXT)', COL_DONE_E: 'Done (EXT)', COL_REM_E: 'Rem (EXT)',
    COL_MP_E: 'Manpower (EXT)', COL_DUR_E: 'Duration (EXT)', COL_MD_E: 'Total Manday (EXT)',
    COL_QTY_I: 'QTY (INT)', COL_DONE_I: 'Done (INT)', COL_REM_I: 'Rem (INT)',
    COL_MP_I: 'Manpower (INT)', COL_DUR_I: 'Duration (INT)', COL_MD_I: 'Total Manday (INT)',
}

# WT04 sub-items
wt04_subs = [
    {'wbs': '1.2.3.1', 'desc': 'WT-04 Schuco Adaptor Profile', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Schuco adaptor profile'},
    {'wbs': '1.2.3.2', 'desc': 'WT-04 Stone Installation', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Stone installation'},
    {'wbs': '1.2.3.3', 'desc': 'WT-04 Glass Installation', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Glass installation'},
    {'wbs': '1.2.3.4', 'desc': 'WT-04 Cocoon-Coordination', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Cocoon coordination'},
    {'wbs': '1.2.3.5', 'desc': 'WT-04 Window Handle Installation + Adjustment', 'bldg': 'E2S', 'scope': 'INT', 'notes': 'Interior - Handle install'},
    {'wbs': '1.2.3.6', 'desc': 'WT-04 Silicone/Gasket Control + Rectification', 'bldg': 'E2S', 'scope': 'INT', 'notes': 'Interior - Silicone/gasket'},
    {'wbs': '1.2.3.7', 'desc': 'WT-04 Interior Membrane Check + Touchup', 'bldg': 'E2S', 'scope': 'INT', 'notes': 'Interior - Membrane check'},
]

# Highlight fill for empty manday
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


def fill_agg_cols(row_arr, wbs_code):
    """Fill aggregation columns for a given WBS code into a row array."""
    if wbs_code not in wbs_agg:
        return False

    a = wbs_agg[wbs_code]
    ext = a['EXT']
    intr = a['INT']
    has_ext = ext['has_data']
    has_int = intr['has_data']

    if has_ext and has_int:
        row_arr[COL_SCOPE] = 'EXT+INT'
    elif has_ext:
        row_arr[COL_SCOPE] = 'EXT'
    elif has_int:
        row_arr[COL_SCOPE] = 'INT'
    else:
        return False

    if has_ext:
        # QTY: if 0 set to 1
        row_arr[COL_QTY_E] = ext['qty'] if ext['qty'] > 0 else 1
        row_arr[COL_DONE_E] = ext['done'] if ext['done'] else None
        row_arr[COL_REM_E] = ext['rem'] if ext['rem'] else None
        row_arr[COL_MP_E] = ext['manpower'] if ext['manpower'] else None
        row_arr[COL_DUR_E] = ext['duration'] if ext['duration'] else None
        row_arr[COL_MD_E] = ext['manday'] if ext['manday'] else None

    if has_int:
        row_arr[COL_QTY_I] = intr['qty'] if intr['qty'] > 0 else 1
        row_arr[COL_DONE_I] = intr['done'] if intr['done'] else None
        row_arr[COL_REM_I] = intr['rem'] if intr['rem'] else None
        row_arr[COL_MP_I] = intr['manpower'] if intr['manpower'] else None
        row_arr[COL_DUR_I] = intr['duration'] if intr['duration'] else None
        row_arr[COL_MD_I] = intr['manday'] if intr['manday'] else None

    return True


def process_e2ns_sheet(ws):
    # Read existing
    rows_val = []
    rows_fmt = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9):
        vals = [c.value for c in row]
        fmts = []
        for c in row:
            fmts.append({
                'font': copy(c.font), 'fill': copy(c.fill),
                'border': copy(c.border), 'alignment': copy(c.alignment),
                'nf': c.number_format,
            })
        rows_val.append(vals)
        rows_fmt.append(fmts)

    new_rows = []
    new_types = []
    new_orig_idx = []
    inserted = 0

    for i, vals in enumerate(rows_val):
        wbs = str(vals[0]) if vals[0] is not None else ''
        extended = list(vals) + [None] * NEW_COLS

        # Header row
        if i == 3:
            for ci, hdr in HEADERS.items():
                extended[ci] = hdr

        # Populate aggregation
        if wbs and wbs != 'WBS':
            fill_agg_cols(extended, wbs)

        new_rows.append(extended)
        new_types.append('orig')
        new_orig_idx.append(i)

        # Insert WT04 subs after 1.2.3
        if wbs == '1.2.3':
            for sub in wt04_subs:
                sr = [None] * TOTAL_COLS
                sr[0] = sub['wbs']
                sr[1] = sub['desc']
                sr[2] = sub['bldg']
                sr[8] = sub['notes']
                sr[COL_SCOPE] = sub['scope']
                fill_agg_cols(sr, sub['wbs'])
                new_rows.append(sr)
                new_types.append('sub')
                new_orig_idx.append(i)
                inserted += 1

        # Insert 1.1.12 after 1.1.11
        elif wbs == '1.1.11':
            sr = [None] * TOTAL_COLS
            sr[0] = '1.1.12'
            sr[1] = 'Cleaning / Finishing / Documentation E2N'
            sr[2] = 'E2N'
            sr[8] = 'Documentation + Testing + Glass replacement'
            fill_agg_cols(sr, '1.1.12')
            new_rows.append(sr)
            new_types.append('sub')
            new_orig_idx.append(i)
            inserted += 1

        # Insert 1.2.10 after 1.2.9
        elif wbs == '1.2.9':
            sr = [None] * TOTAL_COLS
            sr[0] = '1.2.10'
            sr[1] = 'Cleaning / Finishing / Documentation E2S'
            sr[2] = 'E2S'
            sr[8] = 'Documentation + Testing E2S'
            fill_agg_cols(sr, '1.2.10')
            new_rows.append(sr)
            new_types.append('sub')
            new_orig_idx.append(i)
            inserted += 1

    # Unmerge
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        ws.unmerge_cells(str(m))

    # Clear
    for r in range(1, ws.max_row + inserted + 5):
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = 'General'

    # Write
    sub_font = Font(italic=True, size=10, color='444444')
    sub_align = Alignment(horizontal='left', indent=1)
    header_font = Font(bold=True, size=10)

    for r_idx in range(len(new_rows)):
        vals = new_rows[r_idx]
        rtype = new_types[r_idx]
        oidx = new_orig_idx[r_idx]
        excel_row = r_idx + 1
        is_header = (oidx == 3)

        for c in range(TOTAL_COLS):
            cell = ws.cell(row=excel_row, column=c + 1)
            cell.value = vals[c]

            if rtype == 'orig' and c < 9 and oidx < len(rows_fmt):
                fmt = rows_fmt[oidx][c]
                cell.font = fmt['font']
                cell.fill = fmt['fill']
                cell.border = fmt['border']
                cell.alignment = fmt['alignment']
                cell.number_format = fmt['nf']
            elif rtype == 'orig' and c >= 9:
                if is_header:
                    cell.font = header_font
                else:
                    cell.number_format = '#,##0' if c >= 10 else 'General'
            elif rtype == 'sub':
                if c < 9:
                    cell.font = sub_font
                    cell.alignment = sub_align
                else:
                    cell.number_format = '#,##0' if c >= 10 else 'General'

    # Second pass: highlight empty Total Manday cells
    # User wants: if Total Manday is empty for a data row, paint it yellow
    # This catches WBS items with NO Terminplan mapping (no manday data at all)
    highlight_count = 0
    # Identify which rows are section headers (like "1. Facade Installation") vs data rows
    # Section headers typically have short WBS like "1" or "1.1", data rows have "1.1.1" etc.
    # Skip rows 1-4 (title/header rows)
    for r_idx in range(4, len(new_rows)):
        vals = new_rows[r_idx]
        excel_row = r_idx + 1
        wbs_val = vals[0]
        if wbs_val is None or str(wbs_val).strip() == '':
            continue

        wbs_str = str(wbs_val).strip()
        # Skip very top-level section headers (like "1", "1.1", "1.2")
        parts = wbs_str.split('.')
        if len(parts) < 3:
            continue

        # This is a data row (level 3+). Check Total Manday EXT
        md_ext = vals[COL_MD_E]
        if md_ext is None or md_ext == 0 or md_ext == '':
            cell = ws.cell(row=excel_row, column=COL_MD_E + 1)
            cell.fill = YELLOW_FILL
            highlight_count += 1

        # Check Total Manday INT only if scope indicates INT
        scope_val = vals[COL_SCOPE]
        if scope_val and 'INT' in str(scope_val):
            md_int = vals[COL_MD_I]
            if md_int is None or md_int == 0 or md_int == '':
                cell = ws.cell(row=excel_row, column=COL_MD_I + 1)
                cell.fill = YELLOW_FILL
                highlight_count += 1

    print("  E2NS_WBS: %d new rows, %d total, %d cols, %d yellow highlights" % (inserted, len(new_rows), TOTAL_COLS, highlight_count))
    return inserted


e = process_e2ns_sheet(wb['E2NS_WBS'])

# ============================================================
# STEP 5: Change_Log
# ============================================================
ws_log = wb['Change_Log']
next_row = ws_log.max_row + 1
last_num = 0
for r in range(ws_log.max_row, 0, -1):
    v = ws_log.cell(row=r, column=1).value
    if isinstance(v, (int, float)):
        last_num = int(v)
        break
base_num = last_num + 1

changes = [
    ("Cols J-V updated", "Added Scope + QTY/Done/Rem/Manpower/Duration/TotalManday for EXT and INT from Terminplan", "E2NS"),
    ("QTY default", "QTY set to 1 where Terminplan QTY was empty/0", "E2NS"),
    ("Manday highlight", "Yellow highlight on empty Total Manday cells where other data exists", "E2NS"),
]

for i, (sec, desc, pkg) in enumerate(changes):
    r = next_row + i
    ws_log.cell(row=r, column=1, value=base_num + i)
    ws_log.cell(row=r, column=2, value=sec)
    ws_log.cell(row=r, column=3, value=desc)
    ws_log.cell(row=r, column=4, value=pkg)

print("  Change_Log: %d entries" % len(changes))

wb.save(src)
print("\nSaved: %s" % src)
print("DONE!")
