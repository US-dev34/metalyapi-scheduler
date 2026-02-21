import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import shutil
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

src = r"U:\Antigravity\Antigravity\Construction Schedule\WBS_E2NS_D1D2_Restructured.xlsx"
backup = r"U:\Antigravity\Antigravity\Construction Schedule\WBS_E2NS_D1D2_Restructured_BACKUP_v2.xlsx"
shutil.copy2(src, backup)
print("Backup created")

wb = openpyxl.load_workbook(src)

# =============================================
# E2NS Sub-breakdowns
# =============================================
e2ns_subs = {
    '1.1.1': [
        {'wbs': '1.1.1.1', 'desc': 'FT07 Punch Window East Facade', 'bldg': 'E2N', 'notes': 'FT07 punch windows'},
        {'wbs': '1.1.1.2', 'desc': 'FT08 Punch Window East Facade', 'bldg': 'E2N', 'notes': 'FT08 punch windows'},
        {'wbs': '1.1.1.3', 'desc': 'FT11 East Facade', 'bldg': 'E2N', 'notes': 'FT11 scope'},
        {'wbs': '1.1.1.4', 'desc': 'FT15 East Remaining Installation', 'bldg': 'E2N', 'notes': 'FT15 East panels'},
    ],
    '1.1.2': [
        {'wbs': '1.1.2.1', 'desc': 'FT15 E2N North', 'bldg': 'E2N', 'notes': 'FT15 North facade'},
        {'wbs': '1.1.2.2', 'desc': 'FT15 E2N West', 'bldg': 'E2N', 'notes': 'FT15 West facade'},
        {'wbs': '1.1.2.3', 'desc': 'FT15 E2S East', 'bldg': 'E2S', 'notes': 'FT15 E2S East facade'},
        {'wbs': '1.1.2.4', 'desc': 'FT15 E2S South', 'bldg': 'E2S', 'notes': 'FT15 E2S South facade'},
        {'wbs': '1.1.2.5', 'desc': 'FT15 E2S West', 'bldg': 'E2S', 'notes': 'FT15 E2S West facade'},
        {'wbs': '1.1.2.6', 'desc': 'Terminal Cable Canals', 'bldg': 'E2S', 'nta': 'NTA060', 'notes': 'NTA060 Terminal cable canals'},
    ],
    '1.1.3': [
        {'wbs': '1.1.3.1', 'desc': 'Lochfenster E2N', 'bldg': 'E2N', 'notes': 'Punch windows E2N'},
        {'wbs': '1.1.3.2', 'desc': 'Lochfenster E2S', 'bldg': 'E2S', 'notes': 'Punch windows E2S'},
    ],
    '1.1.8': [
        {'wbs': '1.1.8.1', 'desc': 'Attikaabdeckungen E2N', 'bldg': 'E2N', 'notes': 'Parapet capping E2N'},
        {'wbs': '1.1.8.2', 'desc': 'Attikaabdeckungen E2S', 'bldg': 'E2S', 'notes': 'Parapet capping E2S'},
    ],
    '1.2.1': [
        {'wbs': '1.2.1.1', 'desc': 'PR-Fassade FT02', 'bldg': 'E2S', 'notes': 'FT02 curtain wall'},
        {'wbs': '1.2.1.2', 'desc': 'PR-Fassade FT03', 'bldg': 'E2S', 'notes': 'FT03 curtain wall'},
        {'wbs': '1.2.1.3', 'desc': 'PR-Fassade FT10', 'bldg': 'E2S', 'notes': 'FT10 curtain wall'},
        {'wbs': '1.2.1.4', 'desc': 'PR-Fassade FT14', 'bldg': 'E2S', 'notes': 'FT14 curtain wall'},
    ],
    '1.2.2': [
        {'wbs': '1.2.2.1', 'desc': 'PR-Fassade FT05', 'bldg': 'E2S', 'notes': 'FT05 curtain wall'},
        {'wbs': '1.2.2.2', 'desc': 'PR-Fassade FT06', 'bldg': 'E2S', 'notes': 'FT06 curtain wall'},
        {'wbs': '1.2.2.3', 'desc': 'PR-Fassade FT09', 'bldg': 'E2S', 'notes': 'FT09 curtain wall'},
        {'wbs': '1.2.2.4', 'desc': 'Interkom Glass', 'bldg': 'E2S', 'nta': 'NTA063', 'notes': 'NTA063 Interkom glasses'},
    ],
    '1.2.4': [
        {'wbs': '1.2.4.1', 'desc': 'Lochfenster WT01.5', 'bldg': 'E2S', 'notes': 'WT01.5 punch windows'},
        {'wbs': '1.2.4.2', 'desc': 'Lochfenster WT01.6', 'bldg': 'E2S', 'notes': 'WT01.6 punch windows'},
        {'wbs': '1.2.4.3', 'desc': 'Lochfenster WT01.7', 'bldg': 'E2S', 'notes': 'WT01.7 punch windows'},
    ],
}

# =============================================
# D1D2 Sub-breakdowns
# =============================================
d1d2_subs = {
    '2.1.1': [
        {'wbs': '2.1.1.1', 'desc': 'Interior Facade Defects D2 Floors 1-6', 'bldg': 'D2', 'notes': 'D2 lower floors'},
        {'wbs': '2.1.1.2', 'desc': 'Interior Facade Defects D2 Floors 7-12', 'bldg': 'D2', 'notes': 'D2 upper floors'},
    ],
    '2.1.2': [
        {'wbs': '2.1.2.1', 'desc': 'Interior Facade Defects D1 Floors 1-6', 'bldg': 'D1', 'notes': 'D1 lower floors'},
        {'wbs': '2.1.2.2', 'desc': 'Interior Facade Defects D1 Floors 7-12', 'bldg': 'D1', 'notes': 'D1 upper floors'},
    ],
    '2.1.3': [
        {'wbs': '2.1.3.1', 'desc': 'Exterior Facade Defects D1', 'bldg': 'D1', 'notes': 'D1 exterior defects'},
        {'wbs': '2.1.3.2', 'desc': 'Exterior Facade Defects D2', 'bldg': 'D2', 'notes': 'D2 exterior defects'},
    ],
    '2.1.4': [
        {'wbs': '2.1.4.1', 'desc': 'Glass Replacement D1', 'bldg': 'D1', 'notes': 'D1 glass replacement'},
        {'wbs': '2.1.4.2', 'desc': 'Glass Replacement D2', 'bldg': 'D2', 'notes': 'D2 glass replacement'},
    ],
    '2.1.6': [
        {'wbs': '2.1.6.1', 'desc': 'Doors Defects D1', 'bldg': 'D1', 'notes': 'D1 door defects'},
        {'wbs': '2.1.6.2', 'desc': 'Doors Defects D2', 'bldg': 'D2', 'notes': 'D2 door defects'},
    ],
    '2.1.7': [
        {'wbs': '2.1.7.1', 'desc': 'Bird Deterrence D2 Floors 1-6', 'bldg': 'D2', 'notes': 'D2 lower floors'},
        {'wbs': '2.1.7.2', 'desc': 'Bird Deterrence D2 Floors 7-12', 'bldg': 'D2', 'notes': 'D2 upper floors'},
    ],
    '2.1.8': [
        {'wbs': '2.1.8.1', 'desc': 'Bird Deterrence D1 Floors 1-6', 'bldg': 'D1', 'notes': 'D1 lower floors'},
        {'wbs': '2.1.8.2', 'desc': 'Bird Deterrence D1 Floors 7-12', 'bldg': 'D1', 'notes': 'D1 upper floors'},
    ],
    '2.1.9': [
        {'wbs': '2.1.9.1', 'desc': 'Final Facade Cleaning D1', 'bldg': 'D1', 'notes': 'D1 final cleaning'},
        {'wbs': '2.1.9.2', 'desc': 'Final Facade Cleaning D2', 'bldg': 'D2', 'notes': 'D2 final cleaning'},
    ],
}


def process_sheet(ws, subs_dict, label):
    """Read all rows, insert sub-items after matching parents, rewrite sheet."""
    rows_val = []
    rows_fmt = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9):
        vals = [c.value for c in row]
        fmts = []
        for c in row:
            fmts.append({
                'font': copy(c.font),
                'fill': copy(c.fill),
                'border': copy(c.border),
                'alignment': copy(c.alignment),
                'nf': c.number_format,
            })
        rows_val.append(vals)
        rows_fmt.append(fmts)

    new_vals = []
    new_types = []
    new_orig_idx = []
    count = 0

    for i, vals in enumerate(rows_val):
        wbs = str(vals[0]) if vals[0] is not None else ''
        new_vals.append(list(vals))
        new_types.append('orig')
        new_orig_idx.append(i)

        if wbs in subs_dict:
            old_notes = vals[8] or ''
            n = len(subs_dict[wbs])
            tag = "Parent group: %d sub-levels" % n
            if 'sub-level' not in old_notes.lower() and 'sub-group' not in old_notes.lower() and 'Parent group' not in old_notes:
                new_vals[-1][8] = tag if not old_notes else "%s | %s" % (old_notes, tag)

            for sub in subs_dict[wbs]:
                sv = [None]*9
                sv[0] = sub['wbs']
                sv[1] = sub['desc']
                sv[2] = sub.get('bldg')
                sv[3] = sub.get('nta')
                sv[4] = sub.get('status')
                sv[5] = sub.get('budget')
                sv[6] = sub.get('payment')
                sv[7] = sub.get('target_kw')
                sv[8] = sub.get('notes')
                new_vals.append(sv)
                new_types.append('sub')
                new_orig_idx.append(i)
                count += 1

    # Unmerge all merged cells first
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        ws.unmerge_cells(str(m))

    # Clear sheet completely
    for r in range(1, ws.max_row + count + 5):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = 'General'

    sub_font = Font(italic=True, size=10, color='444444')
    sub_align = Alignment(horizontal='left', indent=1)

    for r_idx in range(len(new_vals)):
        vals = new_vals[r_idx]
        rtype = new_types[r_idx]
        oidx = new_orig_idx[r_idx]
        excel_row = r_idx + 1

        for c in range(9):
            cell = ws.cell(row=excel_row, column=c+1)
            cell.value = vals[c]

            if rtype == 'orig':
                fmt = rows_fmt[oidx][c]
                cell.font = fmt['font']
                cell.fill = fmt['fill']
                cell.border = fmt['border']
                cell.alignment = fmt['alignment']
                cell.number_format = fmt['nf']
            else:
                cell.font = sub_font
                cell.alignment = sub_align
                if c == 5:
                    cell.number_format = '#,##0.00'

    print("  %s: %d sub-rows inserted, total rows now %d" % (label, count, len(new_vals)))
    return count


e = process_sheet(wb['E2NS_WBS'], e2ns_subs, 'E2NS_WBS')
d = process_sheet(wb['D1D2_WBS'], d1d2_subs, 'D1D2_WBS')

# =============================================
# Update Change_Log
# =============================================
ws_log = wb['Change_Log']
next_row = ws_log.max_row + 1
base_num = 11

changes = [
    ("1.1.1 -> 4 sub-items", "FT07/FT08/FT11/FT15 East split into 4 facade types", "E2NS"),
    ("1.1.2 -> 6 sub-items", "FT15 E2N N/W + E2S E/S/W + NTA060 Terminal Cable Canals", "E2NS"),
    ("1.1.3 -> 2 sub-items", "Lochfenster split: E2N / E2S", "E2NS"),
    ("1.1.8 -> 2 sub-items", "Attikaabdeckungen split: E2N / E2S", "E2NS"),
    ("1.2.1 -> 4 sub-items", "PR-Fassade FT02/FT03/FT10/FT14 each as sub-item", "E2NS"),
    ("1.2.2 -> 4 sub-items", "PR-Fassade FT05/FT06/FT09 + NTA063 Interkom as sub-items", "E2NS"),
    ("1.2.4 -> 3 sub-items", "Lochfenster WT01.5/WT01.6/WT01.7 each as sub-item", "E2NS"),
    ("2.1.1 -> 2 sub-items", "Interior Defects D2: Floors 1-6 / Floors 7-12", "D1D2"),
    ("2.1.2 -> 2 sub-items", "Interior Defects D1: Floors 1-6 / Floors 7-12", "D1D2"),
    ("2.1.3 -> 2 sub-items", "Exterior Defects split: D1 / D2", "D1D2"),
    ("2.1.4 -> 2 sub-items", "Glass Replacement split: D1 / D2", "D1D2"),
    ("2.1.6 -> 2 sub-items", "Doors Defects split: D1 / D2", "D1D2"),
    ("2.1.7 -> 2 sub-items", "Bird Deterrence D2: Floors 1-6 / Floors 7-12", "D1D2"),
    ("2.1.8 -> 2 sub-items", "Bird Deterrence D1: Floors 1-6 / Floors 7-12", "D1D2"),
    ("2.1.9 -> 2 sub-items", "Final Cleaning split: D1 / D2", "D1D2"),
]

for i, (sec, desc, pkg) in enumerate(changes):
    r = next_row + i
    ws_log.cell(row=r, column=1, value=base_num + i)
    ws_log.cell(row=r, column=2, value=sec)
    ws_log.cell(row=r, column=3, value=desc)
    ws_log.cell(row=r, column=4, value=pkg)

print("  Change_Log: %d entries added" % len(changes))

wb.save(src)
print("\nSaved: %s" % src)
print("DONE!")
