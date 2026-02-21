"""WBS Items router.

Endpoints
---------
GET    /api/v1/wbs/{project_id}/items          WBS tree
POST   /api/v1/wbs/{project_id}/items          Create WBS item
PUT    /api/v1/wbs/{project_id}/items/{id}     Update WBS item
POST   /api/v1/wbs/{project_id}/import         Bulk import from Excel
GET    /api/v1/wbs/{project_id}/export         Export (Phase 4)
"""

import logging
from uuid import UUID

from fastapi import APIRouter, HTTPException, UploadFile, File, status

from backend.models.db import get_db

logger = logging.getLogger(__name__)
from backend.models.schemas import (
    ErrorResponse,
    WBSItemCreate,
    WBSItemResponse,
    WBSItemUpdate,
)
from backend.services.schedule_service import ScheduleService

router = APIRouter(prefix="/api/v1/wbs", tags=["wbs"])
service = ScheduleService()


@router.get("/{project_id}/items", response_model=list[WBSItemResponse])
async def list_wbs_items(project_id: UUID):
    """Return all WBS items for a project, ordered by sort_order."""
    return service.list_wbs_items(project_id)


@router.post(
    "/{project_id}/items",
    response_model=WBSItemResponse,
    status_code=status.HTTP_201_CREATED,
    responses={422: {"model": ErrorResponse}},
)
async def create_wbs_item(project_id: UUID, payload: WBSItemCreate):
    """Create a new WBS item."""
    try:
        return service.create_wbs_item(project_id, payload)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"error": str(exc), "code": "WBS_INVALID_CODE"},
        ) from exc


@router.put(
    "/{project_id}/items/{item_id}",
    response_model=WBSItemResponse,
    responses={404: {"model": ErrorResponse}},
)
async def update_wbs_item(project_id: UUID, item_id: UUID, payload: WBSItemUpdate):
    """Partially update a WBS item."""
    result = service.update_wbs_item(project_id, item_id, payload)
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"WBS item {item_id} not found", "code": "WBS_NOT_FOUND"},
        )
    return result


@router.post("/{project_id}/import")
async def import_wbs(project_id: UUID, file: UploadFile = File(...)):
    """Import WBS items from Excel file."""
    import openpyxl
    from io import BytesIO

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    db = get_db()
    imported = 0
    errors = []

    headers = [cell.value for cell in ws[1]]

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            wbs_code = data.get("wbs_code")
            if not wbs_code:
                continue

            # Resolve parent_code to parent_id
            parent_id = None
            parent_code = data.get("parent_code")
            if parent_code:
                parent_resp = (
                    db.table("wbs_items")
                    .select("id")
                    .eq("project_id", str(project_id))
                    .eq("wbs_code", parent_code)
                    .execute()
                )
                if parent_resp.data:
                    parent_id = parent_resp.data[0]["id"]

            item = {
                "project_id": str(project_id),
                "wbs_code": str(wbs_code),
                "wbs_name": str(data.get("wbs_name", "")),
                "qty": float(data.get("qty", 0) or 0),
                "unit": str(data.get("unit", "pcs") or "pcs"),
                "parent_id": parent_id,
                "level": int(data.get("level", 0) or 0),
                "is_summary": bool(data.get("is_summary", False)),
                "sort_order": row_num,
            }

            db.table("wbs_items").upsert(
                item, on_conflict="project_id,wbs_code"
            ).execute()
            imported += 1
        except Exception as e:
            logger.warning("Import row %d failed: %s", row_num, e)
            errors.append({"row": row_num, "error": str(e)})

    return {"imported": imported, "errors": errors}
