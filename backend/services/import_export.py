"""Import / export service -- Excel file handling via openpyxl.

Provides:
- ``export_to_excel``: serialise a project schedule into a downloadable .xlsx
- ``import_from_excel``: parse an uploaded .xlsx and merge into the project
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any, BinaryIO
from uuid import UUID

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from backend.models.db import get_db

logger = logging.getLogger(__name__)


class ImportExportService:
    """Handles Excel import and export of schedule data."""

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_to_excel(
        self, project_id: UUID
    ) -> tuple[BinaryIO, str]:
        """Generate an .xlsx workbook and return ``(stream, filename)``.

        Sheets created:
        1. **WBS** -- full WBS tree
        2. **Allocations** -- daily allocation cells

        Raises:
            ValueError: If the project does not exist.
        """
        db = get_db()

        project_resp = (
            db.table("projects")
            .select("*")
            .eq("id", str(project_id))
            .execute()
        )
        if not project_resp.data:
            raise ValueError(f"Project {project_id} not found")
        project = project_resp.data[0]

        wbs_items = (
            db.table("wbs_items")
            .select("*")
            .eq("project_id", str(project_id))
            .order("sort_order")
            .execute()
            .data
        )
        # Get WBS item IDs for this project, then fetch allocations
        wbs_ids = [w["id"] for w in wbs_items]
        allocations = []
        if wbs_ids:
            allocations = (
                db.table("daily_allocations")
                .select("*")
                .in_("wbs_item_id", wbs_ids)
                .order("date")
                .execute()
                .data
            )

        wb = Workbook()

        # -- WBS sheet -------------------------------------------------
        ws_wbs = wb.active
        ws_wbs.title = "WBS"  # type: ignore[union-attr]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        wbs_headers = [
            "WBS Code", "WBS Name", "Is Summary", "Qty", "Unit",
            "Level", "Sort Order",
        ]
        for col_idx, header in enumerate(wbs_headers, 1):
            cell = ws_wbs.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, item in enumerate(wbs_items, 2):
            ws_wbs.cell(row=row_idx, column=1, value=item.get("wbs_code", ""))  # type: ignore[union-attr]
            ws_wbs.cell(row=row_idx, column=2, value=item.get("wbs_name", ""))  # type: ignore[union-attr]
            ws_wbs.cell(row=row_idx, column=3, value=item.get("is_summary", False))  # type: ignore[union-attr]
            ws_wbs.cell(row=row_idx, column=4, value=item.get("qty", 0))  # type: ignore[union-attr]
            ws_wbs.cell(row=row_idx, column=5, value=item.get("unit", ""))  # type: ignore[union-attr]
            ws_wbs.cell(row=row_idx, column=6, value=item.get("level", 0))  # type: ignore[union-attr]
            ws_wbs.cell(row=row_idx, column=7, value=item.get("sort_order", 0))  # type: ignore[union-attr]

        # -- Allocations sheet -----------------------------------------
        ws_alloc = wb.create_sheet("Allocations")
        alloc_headers = [
            "WBSItemId", "Date", "PlannedManpower", "ActualManpower", "QtyDone", "Notes",
        ]
        for col_idx, header in enumerate(alloc_headers, 1):
            cell = ws_alloc.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, alloc in enumerate(allocations, 2):
            ws_alloc.cell(row=row_idx, column=1, value=alloc.get("wbs_item_id", ""))
            ws_alloc.cell(row=row_idx, column=2, value=alloc.get("date", ""))
            ws_alloc.cell(row=row_idx, column=3, value=alloc.get("planned_manpower", 0))
            ws_alloc.cell(row=row_idx, column=4, value=alloc.get("actual_manpower", 0))
            ws_alloc.cell(row=row_idx, column=5, value=alloc.get("qty_done", 0))
            ws_alloc.cell(row=row_idx, column=6, value=alloc.get("notes", ""))

        # Serialise to bytes stream
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        safe_name = project["name"].replace(" ", "_")[:40]
        filename = f"{safe_name}_schedule_{date.today().isoformat()}.xlsx"
        return stream, filename

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------

    async def import_from_excel(
        self, project_id: UUID, file: UploadFile
    ) -> dict[str, int]:
        """Parse an uploaded .xlsx and merge rows into the project.

        Returns a summary dict like ``{"wbs_items": 12, "allocations": 48}``.
        """
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
        db = get_db()

        imported_wbs = 0
        imported_alloc = 0

        # -- WBS sheet -------------------------------------------------
        if "WBS" in wb.sheetnames:
            ws = wb["WBS"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or not row[0]:
                    continue
                data: dict[str, Any] = {
                    "project_id": str(project_id),
                    "wbs_code": str(row[0]),
                    "wbs_name": str(row[1]) if len(row) > 1 and row[1] else str(row[0]),
                    "is_summary": row[2] in (True, 1, "True", "true") if len(row) > 2 and row[2] else False,
                    "qty": float(row[3]) if len(row) > 3 and row[3] else 0.0,
                    "unit": str(row[4]) if len(row) > 4 and row[4] else "m2",
                    "level": int(row[5]) if len(row) > 5 and row[5] else 0,
                    "sort_order": int(row[6]) if len(row) > 6 and row[6] else 0,
                }
                db.table("wbs_items").insert(data).execute()
                imported_wbs += 1

        # -- Allocations sheet -----------------------------------------
        if "Allocations" in wb.sheetnames:
            ws = wb["Allocations"]
            # Build a code->id lookup from the project's current WBS items
            wbs_resp = (
                db.table("wbs_items")
                .select("id, wbs_code")
                .eq("project_id", str(project_id))
                .execute()
            )
            code_map = {r["wbs_code"]: r["id"] for r in wbs_resp.data}

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or not row[0]:
                    continue
                wbs_ref = str(row[0])
                wbs_item_id = code_map.get(wbs_ref, wbs_ref)  # fall back to raw id
                alloc_data: dict[str, Any] = {
                    "wbs_item_id": str(wbs_item_id),
                    "date": self._parse_date(row[1]) if len(row) > 1 and row[1] else None,
                    "planned_manpower": float(row[2]) if len(row) > 2 and row[2] else 0.0,
                    "actual_manpower": float(row[3]) if len(row) > 3 and row[3] else 0.0,
                    "qty_done": float(row[4]) if len(row) > 4 and row[4] else 0.0,
                    "notes": str(row[5]) if len(row) > 5 and row[5] else None,
                }
                if alloc_data["date"]:
                    db.table("daily_allocations").upsert(
                        alloc_data, on_conflict="wbs_item_id,date"
                    ).execute()
                    imported_alloc += 1

        wb.close()
        return {"wbs_items": imported_wbs, "allocations": imported_alloc}

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_date(value: Any) -> str | None:
        """Best-effort date parsing from Excel cell values."""
        if value is None:
            return None
        if isinstance(value, (date, datetime)):
            return value.strftime("%Y-%m-%d")
        try:
            return date.fromisoformat(str(value)).isoformat()
        except (ValueError, TypeError):
            return None
