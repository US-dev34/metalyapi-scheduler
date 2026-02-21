import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

src = r"U:\Antigravity\Antigravity\Construction Schedule\DATA\WBS_E2NS_D1D2_Restructured.xlsx"
tp_path = r"U:\Antigravity\Antigravity\Construction Schedule\DATA\Terminplan\260113_DRAFT_E2NS Terminplan.xlsx"
backup = r"U:\Antigravity\Antigravity\Construction Schedule\DATA\WBS_E2NS_D1D2_Restructured_BACKUP_v3.xlsx"

shutil.copy2(src, backup)
print("Backup created:", backup)

# ============================================================
# STEP 1: Read Terminplan data (QTY, Done, Rem per Pos)
# ============================================================
tp_wb = openpyxl.load_workbook(tp_path, data_only=True)
tp_ws = tp_wb['E2NS Terminplan']

# Build dict: terminplan_pos -> {qty, done, rem, desc, group_num (col2), sub1 (col3), sub2 (col4)}
tp_data = {}
for row in tp_ws.iter_rows(min_row=12, max_row=tp_ws.max_row, max_col=20):
    pos = row[4].value  # col E = Pos.
    desc = row[5].value  # col F = Aufgabe
    qty = row[9].value   # col J = QTY
    done = row[10].value  # col K = Done
    rem = row[11].value   # col L = Rem
    grp = row[1].value    # col B = group number
    s1 = row[2].value     # col C = sub1
    s2 = row[3].value     # col D = sub2

    if pos is not None:
        tp_data[str(pos)] = {
            'pos': str(pos), 'desc': str(desc or '').strip(),
            'qty': qty, 'done': done, 'rem': rem,
            'grp': grp, 's1': s1, 's2': s2
        }

print(f"Terminplan: {len(tp_data)} positions loaded")

# ============================================================
# STEP 2: Define FULL MAPPING: Terminplan Pos -> WBS code + Scope (EXT/INT)
# ============================================================
# Format: tp_pos -> (wbs_code, scope, notes)
# For group-level items (x.x.0), we aggregate children QTY/Done/Rem

MAPPING = {
    # === E2N (1.0.0) ===
    # 1.1.0 E2N Outside Windows -> WBS 1.1.3 Lochfenster E2N (parent)
    '1.1.0': ('1.1.3.1', 'EXT', 'E2N Outside Windows - Lochfenster E2N'),
    '1.1.1': ('1.1.3.1', 'EXT', 'Dismantling of Drywall + Al Sheet'),
    '1.1.2': ('1.1.3.1', 'EXT', 'Application of New Membrane'),
    '1.1.3': ('1.1.3.1', 'EXT', 'Installation of Al. Sheet'),
    '1.1.4': ('1.1.3.1', 'EXT', 'Installation of New Drywall'),

    # 1.2.0 E2N Courtyard Windows -> WBS 1.1.1.2 FT08
    '1.2.0': ('1.1.1.2', 'EXT', 'E2N Courtyard Windows - FT08'),
    '1.2.1': ('1.1.1.2', 'EXT', 'Dismantling of Drywall'),
    '1.2.2': ('1.1.1.2', 'EXT', 'Application of New Membrane'),
    '1.2.3': ('1.1.1.2', 'EXT', 'Installation of New Drywall'),

    # 1.3.0 Garage entrance aluminum cladding (north) -> WBS 1.1.2.1 FT15 E2N North
    '1.3.0': ('1.1.2.1', 'EXT', 'Garage entrance aluminum cladding (north)'),

    # 1.4.0 Hormann Door (West) -> WBS 1.1.4 Turen E2N
    '1.4.0': ('1.1.4', 'EXT', 'Hormann Door (West) - parent'),
    '1.4.1': ('1.1.4', 'EXT', 'Steel structure'),
    '1.4.2': ('1.1.4', 'EXT', 'Door'),

    # 1.5.0 FT-07 Glass Rotation -> WBS 1.1.1.1 FT07
    '1.5.0': ('1.1.1.1', 'EXT', 'FT-07 Glass Rotation and Gasket replacement'),

    # 1.6.0 Breuninger facade Bottom Al. -> WBS 1.1.2 FT15 group (user said FT15'e)
    '1.6.0': ('1.1.2', 'EXT', 'Breuninger facade Bottom Al. Finishing Reinforcement'),

    # 1.7.0 RT02 Additional Balustrade -> WBS 1.1.7.3 Additional Steel Balustrade (NTA053)
    '1.7.0': ('1.1.7.3', 'EXT', 'RT02 Additional Balustrade Module (NTA)'),

    # 1.8.0 FT-13 Canopy -> WBS 1.1.5 Main Entrance Canopy E2N
    '1.8.0': ('1.1.5', 'EXT', 'FT-13 Canopy Anodized Al. Sheet + Cap Replacement'),

    # 1.9.0 E2N-E3 Attika Connection -> WBS 1.1.8.1 Attikaabdeckungen E2N
    '1.9.0': ('1.1.8.1', 'EXT', 'E2N-E3 Attika Connection Detail (East+North)'),

    # 1.10.0 E2N East Stone FT15 -> WBS 1.1.1.4 FT15 East
    '1.10.0': ('1.1.1.4', 'EXT', 'E2N East Stone FT15 Connection Detail'),

    # 1.11.0 General Glass replacement -> WBS 1.1.10 Final Facade Cleaning E2N
    '1.11.0': ('1.1.10', 'EXT', 'General Glass replacement'),

    # 1.12.0 FT-18 -> WBS 1.1.6 Lamellenfassade FT18 E2N
    '1.12.0': ('1.1.6', 'EXT', 'FT-18 parent'),
    '1.12.1': ('1.1.6', 'EXT', 'Paint repair and bolt replacement'),
    '1.12.2': ('1.1.6', 'EXT', '1141 Parcel / Passage and parapet installation (NTA)'),

    # 1.13.0 Acoustic Roof FT-18 -> WBS 1.1.9 Acoustic Panel System
    '1.13.0': ('1.1.9', 'EXT', 'Acoustic Roof FT-18 parent'),
    '1.13.1': ('1.1.9.1', 'EXT', 'Installed Steel Drilling + Paint Touch up'),
    '1.13.2': ('1.1.9.2', 'EXT', 'Acoustic Roof Secondary Steel Installation'),
    '1.13.3': ('1.1.9.3', 'EXT', 'Module installation'),

    # 1.14.0 Documentation -> new "Cleaning / Finishing" structure
    '1.14.0': ('1.1.12', 'EXT', 'Documentation + Testing E2N'),
    '1.14.1': ('1.1.12', 'EXT', 'Documentation WT01'),
    '1.14.2': ('1.1.12', 'EXT', 'Documentation FT08'),
    '1.14.3': ('1.1.12', 'EXT', 'Documentation FT07'),

    # === E2S Exterior (2.0.0) ===
    # 2.1.0 FT-02 Zigzak -> WBS 1.2.1.1 PR-Fassade FT02
    '2.1.0': ('1.2.1.1', 'EXT', 'FT-02 Zigzak parent'),
    '2.1.1': ('1.2.1.1', 'EXT', 'Zigzak Mangel (West)'),
    '2.1.2': ('1.2.1.1', 'EXT', 'Zigzak Mangel (South)'),
    '2.1.3': ('1.2.1.1', 'EXT', 'Zigzak Mangel (East)'),

    # 2.2.0 WT-1.5 -> WBS 1.2.4.1 Lochfenster WT01.5
    '2.2.0': ('1.2.4.1', 'EXT', 'WT-1.5 parent'),
    '2.2.1': ('1.2.4.1', 'EXT', 'Mangel (Galvanized sheet end Membran) (West)'),

    # 2.3.0 FT-05 -> WBS 1.2.2.1 PR-Fassade FT05
    '2.3.0': ('1.2.2.1', 'EXT', 'FT-05 parent'),
    '2.3.1': ('1.2.2.1', 'EXT', 'Glass installation'),
    '2.3.2': ('1.2.2.1', 'EXT', 'Membrane control (Big)'),
    '2.3.3': ('1.2.2.1', 'EXT', 'Membrane control (Small)'),

    # 2.4.0 Attika -> WBS 1.1.8.2 Attikaabdeckungen E2S
    '2.4.0': ('1.1.8.2', 'EXT', 'Attika parent'),
    '2.4.1': ('1.1.8.2', 'EXT', '7.OG'),
    '2.4.2': ('1.1.8.2', 'EXT', '5.OG'),
    '2.4.3': ('1.1.8.2', 'EXT', '3.OG'),
    '2.4.3.1': ('1.1.8.2', 'EXT', '3.OG New Design Coordination'),
    '2.4.3.2': ('1.1.8.2', 'EXT', '3.OG Attika (West)'),
    '2.4.3.3': ('1.1.8.2', 'EXT', '3.OG Attika (East)'),

    # 2.5.0 Steel substructure ceramic -> WBS 1.4.1.2 NTA049
    '2.5.0': ('1.4.1.2', 'EXT', 'Steel substructure for ceramic tile installation (NTA)'),

    # 2.6.0 FT-12 Canopy (Pullman) -> WBS 1.2.7 Main Entrance Canopy E2S
    '2.6.0': ('1.2.7', 'EXT', 'FT-12 Canopy (Pullman) parent'),
    '2.6.1': ('1.2.7', 'EXT', 'Steel installation'),
    '2.6.2': ('1.2.7', 'EXT', 'Aluminium Substructure'),
    '2.6.3': ('1.2.7', 'EXT', 'Aluminium Cladding'),

    # 2.7.0 Dorma Sliding door -> WBS 1.2.5 Turen E2S
    '2.7.0': ('1.2.5', 'EXT', 'Dorma Sliding door track installation (East)'),

    # 2.8.0 FT14 Spandrel (NTA) -> WBS 1.4.3.1/1.4.3.2
    '2.8.0': ('1.4.3.1', 'EXT', 'FT14 Additional Spandrel Panel parent'),
    '2.8.1': ('1.4.3.1', 'EXT', 'New Design'),
    '2.8.2': ('1.4.3.1', 'EXT', 'Approval by SOB / NTA Approval by URW'),
    '2.8.3': ('1.4.3.1', 'EXT', 'Procurement + Manufacturing'),
    '2.8.4': ('1.4.3.2', 'EXT', 'Installation'),

    # 2.9.0 Additional Vent Louver -> WBS 1.4.3.3/1.4.3.4
    '2.9.0': ('1.4.3.3', 'EXT', 'Additional Vent backside of Louver parent'),
    '2.9.1': ('1.4.3.3', 'EXT', 'New Design'),
    '2.9.2': ('1.4.3.3', 'EXT', 'Approval by SOB / NTA Approval by URW'),
    '2.9.3': ('1.4.3.3', 'EXT', 'Procurement + Manufacturing'),
    '2.9.4': ('1.4.3.4', 'EXT', 'Installation'),

    # 2.10.0 General Glass replacement -> WBS 1.1.11 Final Facade Cleaning E2S
    '2.10.0': ('1.1.11', 'EXT', 'General Glass replacement'),

    # 2.11.0 Interkom Glasses -> WBS 1.2.2.4 Interkom Glass (NTA063)
    '2.11.0': ('1.2.2.4', 'EXT', 'Interkom Glasses (NTA)'),

    # 2.12.0 FT-18 -> WBS 1.2.6 Lamellenfassade FT18 E2S
    '2.12.0': ('1.2.6', 'EXT', 'FT-18 E2S parent'),
    '2.12.1': ('1.2.6', 'EXT', 'Module Installation'),
    '2.12.2': ('1.2.6', 'EXT', 'Paint repair and bolt replacement'),
    '2.12.3': ('1.2.6', 'EXT', 'Trapezoidal sheet installation'),

    # 2.13.0 Doors -> WBS 1.2.5 Turen E2S
    '2.13.0': ('1.2.5', 'EXT', 'Doors E2S parent'),
    '2.13.1': ('1.2.5', 'EXT', 'Terminal Doors'),
    '2.13.2': ('1.2.5', 'EXT', 'Hotel Doors'),

    # === E2S Inenhof / Courtyard (3.0.0) - EXTERIOR ===
    # 3.1.0 WT-04 -> WBS 1.2.3 Lochfenster WT04 (Exterior part)
    '3.1.0': ('1.2.3', 'EXT', 'WT-04 Exterior parent'),
    '3.1.1': ('1.2.3.1', 'EXT', 'Schuco adaptor profile'),
    '3.1.2': ('1.2.3.2', 'EXT', 'Stone Installation'),
    '3.1.3': ('1.2.3.3', 'EXT', 'Glass installation'),
    '3.1.4': ('1.2.3.4', 'EXT', 'Cocoon-Coordination'),

    # 3.2.0 WT-1.5 -> WBS 1.2.4.1 Lochfenster WT01.5
    '3.2.0': ('1.2.4.1', 'EXT', 'WT-1.5 Inenhof parent'),
    '3.2.1': ('1.2.4.1', 'EXT', 'Mangel (Galvanized sheet end Membran) (West)'),
    '3.2.2': ('1.2.4.1', 'EXT', 'Schuco adaptor profile'),

    # 3.3.0 FT-03 -> WBS 1.2.1.2 PR-Fassade FT03
    '3.3.0': ('1.2.1.2', 'EXT', 'FT-03 Inenhof parent'),
    '3.3.1': ('1.2.1.2', 'EXT', 'Phonotherm structural board + Flashing'),
    '3.3.2': ('1.2.1.2', 'EXT', 'Glass Installation - Pressure + Cover'),
    '3.3.3': ('1.2.1.2', 'EXT', 'Al. sheet under console + steel (East-West)'),
    '3.3.4': ('1.2.1.2', 'EXT', 'Top Finish Attika (North)'),
    '3.3.5': ('1.2.1.2', 'EXT', 'Steel behind stone cladding (South)'),

    # 3.4.0 RT-01 Glass Balustrade -> WBS 1.1.7.1 Stabgelander E2S
    '3.4.0': ('1.1.7.1', 'EXT', 'RT-01 Glass Balustrade parent'),
    '3.4.1': ('1.1.7.1', 'EXT', 'Brackets (NTA)'),
    '3.4.2': ('1.1.7.1', 'EXT', 'Roofing Works'),
    '3.4.3': ('1.1.7.1', 'EXT', 'Steel Construction'),
    '3.4.4': ('1.1.7.1', 'EXT', 'Aluminum Module'),
    '3.4.5': ('1.1.7.1', 'EXT', 'Attika Installation'),
    '3.4.6': ('1.1.7.1', 'EXT', 'Glass installation'),

    # 3.5.0 FT16 A+B -> WBS 1.2.8 Alu Cladding Shafts FT16
    '3.5.0': ('1.2.8', 'EXT', 'FT16 A+B parent'),
    '3.5.1': ('1.2.8', 'EXT', 'Secondary Steel Construction (NTA)'),
    '3.5.2': ('1.2.8', 'EXT', 'Design + Static Calculation'),
    '3.5.3': ('1.2.8', 'EXT', 'Procurement + Manufacturing'),
    '3.5.4': ('1.2.8', 'EXT', 'Installation'),
    '3.5.5': ('1.2.8', 'EXT', 'Brackets'),
    '3.5.6': ('1.2.8', 'EXT', 'Aluminium Profiles'),
    '3.5.7': ('1.2.8', 'EXT', 'Aluminium Cladding'),

    # 3.6.0 FT17 -> WBS 1.2.9 Aluminium Soffit FT17
    '3.6.0': ('1.2.9', 'EXT', 'FT17 parent'),
    '3.6.1': ('1.2.9', 'EXT', 'MEP + Other Works Completion'),
    '3.6.2': ('1.2.9', 'EXT', 'Brackets'),
    '3.6.3': ('1.2.9', 'EXT', 'Aluminium Profiles'),
    '3.6.4': ('1.2.9', 'EXT', 'Aluminium Cladding'),

    # === E2S Interior (4.0.0) ===
    # 4.1.0 FT15 Terminal Cable Canals -> WBS 1.1.2.6 Terminal Cable Canals (NTA060)
    '4.1.0': ('1.1.2.6', 'INT', 'FT15 Terminal Cable Canals parent'),
    '4.1.1': ('1.1.2.6', 'INT', 'New Design'),
    '4.1.2': ('1.1.2.6', 'INT', 'Approval by SOB / NTA Approval by URW'),
    '4.1.3': ('1.1.2.6', 'INT', 'Procurement + Manufacturing'),
    '4.1.4': ('1.1.2.6', 'INT', 'Installation'),

    # 4.2.0 WT-04 Interior -> WBS 1.2.3 Lochfenster WT04 (Interior sub-items)
    '4.2.0': ('1.2.3', 'INT', 'WT-04 Interior parent'),
    '4.2.1': ('1.2.3.5', 'INT', 'Window Handle Installation + Adjustment'),
    '4.2.2': ('1.2.3.6', 'INT', 'Silicone/Gasket Control + Rectification'),
    '4.2.3': ('1.2.3.7', 'INT', 'Interior Membrane Check + Touchup'),

    # 4.3.0 WT-1.5 Interior -> WBS 1.2.4.1 Lochfenster WT01.5
    '4.3.0': ('1.2.4.1', 'INT', 'WT-1.5 Interior parent'),
    '4.3.1': ('1.2.4.1', 'INT', 'Window Handle Installation + Adjustment'),
    '4.3.2': ('1.2.4.1', 'INT', 'Silicone/Gasket Control + Rectification'),
    '4.3.3': ('1.2.4.1', 'INT', 'Interior Membrane Check + Touchup'),

    # 4.4.0 FT-02 Zigzak Interior -> WBS 1.2.1.1 PR-Fassade FT02
    '4.4.0': ('1.2.1.1', 'INT', 'FT-02 Zigzak Interior parent'),
    '4.4.1': ('1.2.1.1', 'INT', 'Window Handle Installation + Adjustment'),
    '4.4.2': ('1.2.1.1', 'INT', 'Silicone/Gasket Control + Rectification'),
    '4.4.3': ('1.2.1.1', 'INT', 'Interior Membrane Check + Touchup'),

    # 4.5.0 FT-14 Interior -> WBS 1.2.1.4 PR-Fassade FT14
    '4.5.0': ('1.2.1.4', 'INT', 'FT-14 Interior parent'),
    '4.5.1': ('1.2.1.4', 'INT', 'Aluminum sheet installation'),
    '4.5.2': ('1.2.1.4', 'INT', 'Window Handle Installation + Adjustment'),

    # 4.6.0 FT-05 Interior -> WBS 1.2.2.1 PR-Fassade FT05
    '4.6.0': ('1.2.2.1', 'INT', 'FT-05 Interior parent'),
    '4.6.1': ('1.2.2.1', 'INT', 'Membrane Control + Rectification'),

    # 4.7.0 FT-03 Interior -> WBS 1.2.1.2 PR-Fassade FT03
    '4.7.0': ('1.2.1.2', 'INT', 'FT-03 Interior parent'),
    '4.7.1': ('1.2.1.2', 'INT', 'Aluminum sheet installation'),
    '4.7.2': ('1.2.1.2', 'INT', 'Window Handle Installation + Adjustment'),
    '4.7.3': ('1.2.1.2', 'INT', 'Paint - Touch up'),

    # 4.8.0 WT-1.6 Interior -> WBS 1.2.4.2 Lochfenster WT01.6
    '4.8.0': ('1.2.4.2', 'INT', 'WT-1.6 Interior parent'),
    '4.8.1': ('1.2.4.2', 'INT', 'Membrane Control + Rectification'),

    # 4.9.0 FT-06 Pool facade -> WBS 1.2.2.2 PR-Fassade FT06
    '4.9.0': ('1.2.2.2', 'INT', 'FT-06 Pool facade parent'),
    '4.9.1': ('1.2.2.2', 'INT', 'Mock-up on Site'),
    '4.9.2': ('1.2.2.2', 'INT', 'Approval by SOB/DSI/URW'),
    '4.9.3': ('1.2.2.2', 'INT', 'Procurement + Manufacturing'),
    '4.9.4': ('1.2.2.2', 'INT', 'Installation'),

    # 4.10.0 FT03 Acoustic Reinforcement -> WBS 1.2.1.2 PR-Fassade FT03 (Interior)
    '4.10.0': ('1.2.1.2', 'INT', 'FT03 Acoustic Reinforcement parent'),
    '4.10.1': ('1.2.1.2', 'INT', 'Mock-up on Site'),
    '4.10.2': ('1.2.1.2', 'INT', 'Acoustic Testing'),
    '4.10.3': ('1.2.1.2', 'INT', 'Approval by SOB/DSI/URW'),
    '4.10.4': ('1.2.1.2', 'INT', 'Procurement + Manufacturing'),
    '4.10.5': ('1.2.1.2', 'INT', 'Installation'),

    # 4.11.0 Documentation + Testing -> 1.1.12 or parent Cleaning/Finishing
    '4.11.0': ('1.2.10', 'INT', 'Documentation + Testing E2S'),
    '4.12.0': (None, None, 'Warehouse + Supervisor - skip'),
}

# ============================================================
# STEP 3: Aggregate QTY/Done/Rem per WBS code, split by EXT/INT
# ============================================================
# Structure: wbs_code -> { 'EXT': {qty, done, rem, activities:[]}, 'INT': {qty, done, rem, activities:[]} }
wbs_agg = {}

for tp_pos, (wbs_code, scope, note) in MAPPING.items():
    if wbs_code is None:
        continue
    tp = tp_data.get(tp_pos)
    if tp is None:
        continue

    if wbs_code not in wbs_agg:
        wbs_agg[wbs_code] = {
            'EXT': {'qty': 0, 'done': 0, 'rem': 0, 'activities': []},
            'INT': {'qty': 0, 'done': 0, 'rem': 0, 'activities': []}
        }

    s = scope or 'EXT'
    q = tp['qty'] if tp['qty'] is not None and isinstance(tp['qty'], (int, float)) else 0
    d = tp['done'] if tp['done'] is not None and isinstance(tp['done'], (int, float)) else 0
    r = tp['rem'] if tp['rem'] is not None and isinstance(tp['rem'], (int, float)) else 0

    # Only add leaf-level items (sub2 != 0 means it's a task, not a group header)
    # But also add group-level items that have their own QTY (like 1.3.0, 1.5.0 etc)
    if q > 0 or d > 0 or r > 0:
        wbs_agg[wbs_code][s]['qty'] += q
        wbs_agg[wbs_code][s]['done'] += d
        wbs_agg[wbs_code][s]['rem'] += r
        wbs_agg[wbs_code][s]['activities'].append(f"{tp_pos}: {note}")

print(f"WBS aggregation: {len(wbs_agg)} WBS codes with data")

# ============================================================
# STEP 4: Load WBS file and update
# ============================================================
wb = openpyxl.load_workbook(src)

# We need to:
# 1. Add columns: Scope (EXT/INT), QTY_EXT, Done_EXT, Rem_EXT, QTY_INT, Done_INT, Rem_INT
# 2. Add new WBS items: 1.2.3.1-7 (WT04 sub-breakdown), 1.1.12 (Documentation E2N), 1.2.10 (Documentation E2S)
# 3. Populate data from Terminplan

# First, let's define new sub-items to add
new_wbs_items = {
    # WT-04 sub-breakdown (user confirmed this structure)
    '1.2.3': [
        {'wbs': '1.2.3.1', 'desc': 'WT-04 Schuco Adaptor Profile', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Schuco adaptor profile'},
        {'wbs': '1.2.3.2', 'desc': 'WT-04 Stone Installation', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Stone installation'},
        {'wbs': '1.2.3.3', 'desc': 'WT-04 Glass Installation', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Glass installation'},
        {'wbs': '1.2.3.4', 'desc': 'WT-04 Cocoon-Coordination', 'bldg': 'E2S', 'scope': 'EXT', 'notes': 'Exterior - Cocoon coordination (depends on them)'},
        {'wbs': '1.2.3.5', 'desc': 'WT-04 Window Handle Installation + Adjustment', 'bldg': 'E2S', 'scope': 'INT', 'notes': 'Interior - Handle install + adjustment'},
        {'wbs': '1.2.3.6', 'desc': 'WT-04 Silicone/Gasket Control + Rectification', 'bldg': 'E2S', 'scope': 'INT', 'notes': 'Interior - Silicone/gasket control'},
        {'wbs': '1.2.3.7', 'desc': 'WT-04 Interior Membrane Check + Touchup', 'bldg': 'E2S', 'scope': 'INT', 'notes': 'Interior - Membrane check + touchup'},
    ],
}

# New standalone items to add at the end of sections
new_standalone = [
    # 1.1.12 Cleaning / Finishing / Documentation E2N
    {'wbs': '1.1.12', 'desc': 'Cleaning / Finishing / Documentation E2N', 'bldg': 'E2N', 'scope': 'EXT',
     'notes': 'Documentation + Testing + Glass replacement', 'after_section': '1.1'},
    # 1.2.10 Cleaning / Finishing / Documentation E2S
    {'wbs': '1.2.10', 'desc': 'Cleaning / Finishing / Documentation E2S', 'bldg': 'E2S', 'scope': 'INT',
     'notes': 'Documentation + Testing E2S Interior', 'after_section': '1.2'},
]


def process_e2ns_sheet(ws):
    """Process E2NS_WBS sheet: add columns, new rows, populate QTY/Done/Rem"""

    # Read existing data
    rows_val = []
    rows_fmt = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9):
        vals = [c.value for c in row]
        fmts = []
        for c in row:
            fmts.append({
                'font': copy(c.font),
                'fill': copy(c.fill),
                'border': copy(c.border),
                'alignment': copy(c.alignment),
                'nf': c.number_format,
            })
        rows_val.append(vals)
        rows_fmt.append(fmts)

    # Build new row list with:
    # - Original 9 cols (A-I) + new cols: J=Scope, K=QTY_EXT, L=Done_EXT, M=Rem_EXT, N=QTY_INT, O=Done_INT, P=Rem_INT
    # - Insert WT04 sub-items after 1.2.3
    # - Insert 1.1.12 after 1.1.11
    # - Insert 1.2.10 after 1.2.9

    NEW_COLS = 7  # Scope, QTY_EXT, Done_EXT, Rem_EXT, QTY_INT, Done_INT, Rem_INT
    total_cols = 9 + NEW_COLS  # 16 cols total

    new_rows = []
    new_types = []
    new_orig_idx = []
    inserted = 0

    for i, vals in enumerate(rows_val):
        wbs = str(vals[0]) if vals[0] is not None else ''

        # Extend to 16 cols
        extended = list(vals) + [None] * NEW_COLS

        # Add header row labels
        if i == 3:  # Header row (row 4)
            extended[9] = 'Scope'
            extended[10] = 'QTY (EXT)'
            extended[11] = 'Done (EXT)'
            extended[12] = 'Rem (EXT)'
            extended[13] = 'QTY (INT)'
            extended[14] = 'Done (INT)'
            extended[15] = 'Rem (INT)'

        # Populate QTY/Done/Rem from aggregation
        if wbs in wbs_agg:
            agg = wbs_agg[wbs]
            ext = agg['EXT']
            intr = agg['INT']

            # Determine scope
            has_ext = ext['qty'] > 0 or ext['done'] > 0 or ext['rem'] > 0
            has_int = intr['qty'] > 0 or intr['done'] > 0 or intr['rem'] > 0
            if has_ext and has_int:
                extended[9] = 'EXT+INT'
            elif has_ext:
                extended[9] = 'EXT'
            elif has_int:
                extended[9] = 'INT'

            if has_ext:
                extended[10] = ext['qty'] if ext['qty'] else None
                extended[11] = ext['done'] if ext['done'] else None
                extended[12] = ext['rem'] if ext['rem'] else None
            if has_int:
                extended[13] = intr['qty'] if intr['qty'] else None
                extended[14] = intr['done'] if intr['done'] else None
                extended[15] = intr['rem'] if intr['rem'] else None

        new_rows.append(extended)
        new_types.append('orig')
        new_orig_idx.append(i)

        # Insert WT04 sub-items after 1.2.3
        if wbs == '1.2.3':
            for sub in new_wbs_items['1.2.3']:
                sub_row = [None] * total_cols
                sub_row[0] = sub['wbs']
                sub_row[1] = sub['desc']
                sub_row[2] = sub['bldg']
                sub_row[8] = sub['notes']
                sub_row[9] = sub['scope']

                # Populate from aggregation
                swbs = sub['wbs']
                if swbs in wbs_agg:
                    a = wbs_agg[swbs]
                    if sub['scope'] == 'EXT':
                        sub_row[10] = a['EXT']['qty'] or None
                        sub_row[11] = a['EXT']['done'] or None
                        sub_row[12] = a['EXT']['rem'] or None
                    else:
                        sub_row[13] = a['INT']['qty'] or None
                        sub_row[14] = a['INT']['done'] or None
                        sub_row[15] = a['INT']['rem'] or None

                new_rows.append(sub_row)
                new_types.append('sub')
                new_orig_idx.append(i)
                inserted += 1

        # Insert 1.1.12 after 1.1.11
        elif wbs == '1.1.11':
            sub_row = [None] * total_cols
            sub_row[0] = '1.1.12'
            sub_row[1] = 'Cleaning / Finishing / Documentation E2N'
            sub_row[2] = 'E2N'
            sub_row[8] = 'Documentation + Testing + Glass replacement'
            sub_row[9] = 'EXT'
            if '1.1.12' in wbs_agg:
                a = wbs_agg['1.1.12']
                sub_row[10] = a['EXT']['qty'] or None
                sub_row[11] = a['EXT']['done'] or None
                sub_row[12] = a['EXT']['rem'] or None
            new_rows.append(sub_row)
            new_types.append('sub')
            new_orig_idx.append(i)
            inserted += 1

        # Insert 1.2.10 after 1.2.9
        elif wbs == '1.2.9':
            sub_row = [None] * total_cols
            sub_row[0] = '1.2.10'
            sub_row[1] = 'Cleaning / Finishing / Documentation E2S'
            sub_row[2] = 'E2S'
            sub_row[8] = 'Documentation + Testing E2S'
            sub_row[9] = 'INT'
            if '1.2.10' in wbs_agg:
                a = wbs_agg['1.2.10']
                sub_row[13] = a['INT']['qty'] or None
                sub_row[14] = a['INT']['done'] or None
                sub_row[15] = a['INT']['rem'] or None
            new_rows.append(sub_row)
            new_types.append('sub')
            new_orig_idx.append(i)
            inserted += 1

    # Unmerge all merged cells
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        ws.unmerge_cells(str(m))

    # Clear sheet
    for r in range(1, ws.max_row + inserted + 5):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = 'General'

    # Write new data
    sub_font = Font(italic=True, size=10, color='444444')
    sub_align = Alignment(horizontal='left', indent=1)
    header_font = Font(bold=True, size=10)

    for r_idx in range(len(new_rows)):
        vals = new_rows[r_idx]
        rtype = new_types[r_idx]
        oidx = new_orig_idx[r_idx]
        excel_row = r_idx + 1

        for c in range(min(len(vals), total_cols)):
            cell = ws.cell(row=excel_row, column=c + 1)
            cell.value = vals[c]

            if rtype == 'orig' and c < 9 and oidx < len(rows_fmt):
                fmt = rows_fmt[oidx][c]
                cell.font = fmt['font']
                cell.fill = fmt['fill']
                cell.border = fmt['border']
                cell.alignment = fmt['alignment']
                cell.number_format = fmt['nf']
            elif rtype == 'orig' and c >= 9:
                # New columns for original rows
                if oidx == 3:  # header row
                    cell.font = header_font
                else:
                    cell.number_format = '#,##0' if c >= 10 else 'General'
            elif rtype == 'sub':
                cell.font = sub_font
                cell.alignment = sub_align
                if c >= 10:
                    cell.number_format = '#,##0'

    print(f"  E2NS_WBS: {inserted} new rows, {len(new_rows)} total rows, {total_cols} columns")
    return inserted


e = process_e2ns_sheet(wb['E2NS_WBS'])

# ============================================================
# STEP 5: Update Change_Log
# ============================================================
ws_log = wb['Change_Log']
next_row = ws_log.max_row + 1
base_num = ws_log.cell(row=ws_log.max_row, column=1).value
if not isinstance(base_num, (int, float)):
    base_num = 25
base_num = int(base_num) + 1

changes = [
    ("1.2.3 -> 7 sub-items", "WT-04 split: EXT (Schuco/Stone/Glass/Cocoon) + INT (Handle/Silicone/Membrane)", "E2NS"),
    ("1.1.12 added", "Cleaning / Finishing / Documentation E2N (Glass replacement + Testing)", "E2NS"),
    ("1.2.10 added", "Cleaning / Finishing / Documentation E2S (Interior Testing)", "E2NS"),
    ("Cols J-P added", "Scope (EXT/INT) + QTY/Done/Rem from Terminplan mapped to WBS", "E2NS"),
]

for i, (sec, desc, pkg) in enumerate(changes):
    r = next_row + i
    ws_log.cell(row=r, column=1, value=base_num + i)
    ws_log.cell(row=r, column=2, value=sec)
    ws_log.cell(row=r, column=3, value=desc)
    ws_log.cell(row=r, column=4, value=pkg)

print(f"  Change_Log: {len(changes)} entries added")

# ============================================================
# SAVE
# ============================================================
wb.save(src)
print(f"\nSaved: {src}")
print("DONE!")
