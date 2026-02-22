"""Rebuild WBS hierarchy using Payment Schedule as base, Terminplan as sub-items.

Payment Schedule = family grouping (L0: sections, L1: payment items)
Terminplan = leaf items with manpower data (L2 under payment items)

The "Weekly Detailed Follow-Up" sheet shows which Terminplan items belong
to each payment item via sub-codes like 1.1.1, 2.1.7, NTA040, etc.
"""

import uuid
import json
import re
import openpyxl
from datetime import datetime

PROJECT_ID = "5f0fc90a-00b7-4cf7-aaba-22dde8118fa1"
NAMESPACE = uuid.UUID(PROJECT_ID)

def make_uuid(code):
    """Deterministic UUID from code string."""
    return str(uuid.uuid5(NAMESPACE, code))

def make_alloc_uuid(pos_code, date_str):
    return str(uuid.uuid5(NAMESPACE, f"alloc:{pos_code}:{date_str}"))

# ─────────────────────────────────────────────────────────
# 1. Read Payment Schedule
# ─────────────────────────────────────────────────────────
print("=== Reading Payment Schedule ===")
ps_wb = openpyxl.load_workbook(
    r"u:\Antigravity\Antigravity\Construction Schedule\DATA\260206_Copy of 260119_Rest to Execute  Payments_URW_TS_MYA Commented.xlsx",
    data_only=True
)

# Read E2NS sheet for main structure
ps = ps_wb["E2NS"]
payment_items = []
for row in range(1, ps.max_row + 1):
    code = ps.cell(row=row, column=1).value
    desc = ps.cell(row=row, column=2).value
    area = ps.cell(row=row, column=3).value
    status = ps.cell(row=row, column=4).value
    target_cw = ps.cell(row=row, column=5).value
    amount = ps.cell(row=row, column=6).value
    if code and desc:
        payment_items.append({
            "code": str(code).strip(),
            "desc": str(desc).strip(),
            "area": str(area).strip() if area else "",
            "status": str(status).strip() if status else "",
            "target_cw": target_cw,
            "amount": float(amount) if amount else 0,
            "row": row,
        })

print(f"  Payment items from E2NS sheet: {len(payment_items)}")
for p in payment_items:
    print(f"    {p['code']:8s} | {p['desc'][:60]:60s} | {p['area']:6s} | {p['status']:6s} | EUR {p['amount']:>12.2f}")

# Read Weekly Detailed Follow-Up for sub-item mapping
wdf = ps_wb["Weekly Detailed Follow-Up"]
mapping = {}  # payment_code -> [sub_codes]
current_payment_code = None

# Payment codes are XX.XX format (e.g., 02.11, 04.12, 05.14)
PAYMENT_CODE_RE = re.compile(r'^\d{2}\.\d{2}$')
# Sub-item arrow character: Unicode U+21B3 (downwards arrow with tip rightwards)
SUB_ARROW = '\u21b3'

for row in range(1, wdf.max_row + 1):
    code = wdf.cell(row=row, column=1).value
    desc = wdf.cell(row=row, column=2).value
    if not desc:
        continue
    desc_str = str(desc).strip()
    code_str = str(code).strip() if code else ""

    # Sub-items have description starting with arrow character
    if SUB_ARROW in desc_str and current_payment_code:
        sub_code = code_str
        if sub_code:
            mapping[current_payment_code].append(sub_code)
    # Payment items have codes like 02.11, 04.12, etc.
    elif PAYMENT_CODE_RE.match(code_str):
        current_payment_code = code_str
        if current_payment_code not in mapping:
            mapping[current_payment_code] = []

total_mapped = sum(len(v) for v in mapping.values())
print(f"\n  Payment->Terminplan mapping from Weekly Detailed Follow-Up ({total_mapped} sub-items):")
for pc, subs in sorted(mapping.items()):
    if subs:
        print(f"    {pc}: {subs}")

# ─────────────────────────────────────────────────────────
# 2. Read Terminplan
# ─────────────────────────────────────────────────────────
print("\n=== Reading Terminplan ===")
tp_wb = openpyxl.load_workbook(
    r"u:\Antigravity\Antigravity\Construction Schedule\DATA\Terminplan\260113_DRAFT_E2NS Terminplan.xlsx",
    data_only=True
)
tp = tp_wb["E2NS Terminplan"]

terminplan_items = []
for row in range(11, 183):
    pos = tp.cell(row=row, column=5).value
    task = tp.cell(row=row, column=6).value
    responsible = tp.cell(row=row, column=8).value
    group = tp.cell(row=row, column=9).value
    qty = tp.cell(row=row, column=10).value
    done = tp.cell(row=row, column=11).value
    manpower = tp.cell(row=row, column=13).value
    duration = tp.cell(row=row, column=14).value

    if not pos:
        continue

    pos = str(pos).strip()
    task_str = str(task).strip() if task else pos

    terminplan_items.append({
        "pos": pos,
        "name": task_str,
        "responsible": str(responsible).strip() if responsible else "",
        "group": str(group).strip() if group else "",
        "qty": float(qty) if qty else 0,
        "done": float(done) if done else 0,
        "manpower": float(manpower) if manpower else 0,
        "duration": float(duration) if duration else 0,
        "row": row,
    })

print(f"  Terminplan items: {len(terminplan_items)}")

# ─────────────────────────────────────────────────────────
# 3. Build WBS hierarchy
# ─────────────────────────────────────────────────────────
print("\n=== Building WBS Hierarchy ===")

# Build reverse mapping: terminplan_pos -> payment_code
tp_to_payment = {}
for pc, subs in mapping.items():
    for sub in subs:
        tp_to_payment[sub] = pc

# Assign each terminplan item to a payment item
tp_assignments = {}  # tp_pos -> payment_code

# First pass: use explicit mapping from Weekly Detailed Follow-Up
for tp_item in terminplan_items:
    pos = tp_item["pos"]
    if pos in tp_to_payment:
        tp_assignments[pos] = tp_to_payment[pos]

# Second pass: children inherit parent's assignment (x.Y.Z inherits from x.Y.0)
for tp_item in terminplan_items:
    pos = tp_item["pos"]
    if pos in tp_assignments:
        continue
    parts = pos.split(".")
    if len(parts) >= 3:
        parent_pos = f"{parts[0]}.{parts[1]}.0"
        if parent_pos != pos and parent_pos in tp_to_payment:
            tp_assignments[pos] = tp_to_payment[parent_pos]
    if len(parts) == 4 and pos not in tp_assignments:
        parent_pos = ".".join(parts[:3])
        if parent_pos in tp_to_payment:
            tp_assignments[pos] = tp_to_payment[parent_pos]

print(f"\n  Assigned {len(tp_assignments)} of {len(terminplan_items)} terminplan items via mapping")

# ─────────────────────────────────────────────────────────
# 4. Build the complete WBS tree
# ─────────────────────────────────────────────────────────
print("\n=== Building Complete WBS Tree ===")

# L0 sections from Payment Schedule + which Terminplan prefixes belong to each
sections = [
    {"code": "E2N", "name": "E2N - North Facade", "ps_prefix": "02", "tp_prefixes": ["1"]},
    {"code": "E2N-NTA", "name": "E2N - Additional Works (NTA)", "ps_prefix": "03", "tp_prefixes": []},
    {"code": "E2S", "name": "E2S - South Facade", "ps_prefix": "04", "tp_prefixes": ["2", "3", "4"]},
    {"code": "E2S-NTA", "name": "E2S - Additional Works (NTA)", "ps_prefix": "05", "tp_prefixes": []},
]

wbs_rows = []
assigned_tp_codes = set()
sort_idx = 0

def add_tp_children(parent_id, tp_items, level):
    """Add terminplan items as children, grouping by Terminplan parent (x.Y.0)."""
    global sort_idx
    # Separate into groups (x.Y.0 parents) and their children
    groups = {}  # parent_pos -> [children]
    standalone = []
    for item in tp_items:
        parts = item["pos"].split(".")
        if len(parts) >= 3 and parts[-1] != "0":
            parent_pos = f"{parts[0]}.{parts[1]}.0"
            groups.setdefault(parent_pos, []).append(item)
        elif len(parts) == 4 and parts[2] != "0":
            parent_pos = ".".join(parts[:3])
            groups.setdefault(parent_pos, []).append(item)
        else:
            standalone.append(item)

    # Add groups: parent as L1 summary, children as L2 leaves
    for parent_pos in sorted(groups.keys()):
        children = groups[parent_pos]
        if parent_pos in assigned_tp_codes:
            continue
        parent_tp = next((t for t in terminplan_items if t["pos"] == parent_pos), None)
        parent_name = parent_tp["name"] if parent_tp else parent_pos

        group_id = make_uuid(parent_pos)
        sort_idx += 1
        assigned_tp_codes.add(parent_pos)
        wbs_rows.append({
            "id": group_id, "parent_id": parent_id,
            "wbs_code": parent_pos, "wbs_name": parent_name,
            "qty": 0, "unit": "pcs", "sort_order": sort_idx * 100,
            "level": level, "is_summary": True,
        })
        for child in sorted(children, key=lambda c: c["pos"]):
            if child["pos"] in assigned_tp_codes:
                continue
            child_id = make_uuid(child["pos"])
            sort_idx += 1
            assigned_tp_codes.add(child["pos"])
            wbs_rows.append({
                "id": child_id, "parent_id": group_id,
                "wbs_code": child["pos"], "wbs_name": child["name"],
                "qty": child["qty"], "unit": "pcs", "sort_order": sort_idx * 100,
                "level": level + 1, "is_summary": False,
            })

    # Add standalone items (x.Y.0 with no unassigned children)
    for item in sorted(standalone, key=lambda s: s["pos"]):
        if item["pos"] in assigned_tp_codes:
            continue
        if item["pos"].endswith(".0.0"):
            continue  # Skip section headers (1.0.0, 2.0.0, etc.)
        item_id = make_uuid(item["pos"])
        sort_idx += 1
        parts = item["pos"].split(".")
        is_parent = len(parts) >= 3 and parts[-1] == "0"
        assigned_tp_codes.add(item["pos"])
        wbs_rows.append({
            "id": item_id, "parent_id": parent_id,
            "wbs_code": item["pos"], "wbs_name": item["name"],
            "qty": item["qty"], "unit": "pcs", "sort_order": sort_idx * 100,
            "level": level, "is_summary": is_parent,
        })


for section in sections:
    section_id = make_uuid(f"section:{section['code']}")
    sort_idx += 1
    wbs_rows.append({
        "id": section_id, "parent_id": None,
        "wbs_code": section["code"], "wbs_name": section["name"],
        "qty": 0, "unit": "pcs", "sort_order": sort_idx * 100,
        "level": 0, "is_summary": True,
    })

    # Get payment items for this section
    section_payments = [p for p in payment_items if p["code"].startswith(section["ps_prefix"] + ".")]

    for pi in section_payments:
        if pi["code"].endswith(".10"):
            continue  # Skip section header rows

        pi_id = make_uuid(f"payment:{pi['code']}")
        sort_idx += 1

        wbs_rows.append({
            "id": pi_id, "parent_id": section_id,
            "wbs_code": pi["code"], "wbs_name": pi["desc"],
            "qty": 0, "unit": "EUR", "sort_order": sort_idx * 100,
            "level": 1, "is_summary": True,
        })

        # Add terminplan items mapped to this payment item
        mapped_children = [t for t in terminplan_items if tp_assignments.get(t["pos"]) == pi["code"]]
        for child in sorted(mapped_children, key=lambda c: c["pos"]):
            if child["pos"] in assigned_tp_codes:
                continue
            child_id = make_uuid(child["pos"])
            sort_idx += 1
            assigned_tp_codes.add(child["pos"])
            wbs_rows.append({
                "id": child_id, "parent_id": pi_id,
                "wbs_code": child["pos"], "wbs_name": child["name"],
                "qty": child["qty"], "unit": "pcs", "sort_order": sort_idx * 100,
                "level": 2, "is_summary": False,
            })

    # After all payment items, add unassigned terminplan items for this section's prefixes
    for tp_prefix in section.get("tp_prefixes", []):
        unassigned_for_prefix = [
            t for t in terminplan_items
            if t["pos"].split(".")[0] == tp_prefix and t["pos"] not in assigned_tp_codes
        ]
        if unassigned_for_prefix:
            add_tp_children(section_id, unassigned_for_prefix, level=1)

# ─────────────────────────────────────────────────────────
# 5. Summary
# ─────────────────────────────────────────────────────────
print(f"\nTotal WBS rows: {len(wbs_rows)}")
level_counts = {}
for r in wbs_rows:
    level_counts[r["level"]] = level_counts.get(r["level"], 0) + 1
print(f"Level breakdown: {level_counts}")

print("\nHierarchy preview:")
for r in wbs_rows:
    indent = "  " * r["level"]
    summary = " [SUMMARY]" if r["is_summary"] else ""
    qty_str = f" qty={r['qty']}" if r["qty"] > 0 else ""
    print(f"  {indent}{r['wbs_code']:12s} {r['wbs_name'][:55]}{summary}{qty_str}")

# ─────────────────────────────────────────────────────────
# 6. Generate SQL
# ─────────────────────────────────────────────────────────
print("\n=== Generating SQL ===")

sql_lines = []

# Step 1: Drop FK constraints to prevent CASCADE deletes
sql_lines.append("-- Temporarily drop FK constraints to prevent cascade deletion of allocations")
sql_lines.append("ALTER TABLE daily_allocations DROP CONSTRAINT IF EXISTS daily_allocations_wbs_item_id_fkey;")
sql_lines.append("ALTER TABLE baseline_snapshots DROP CONSTRAINT IF EXISTS baseline_snapshots_wbs_item_id_fkey;")
sql_lines.append("ALTER TABLE ai_forecasts DROP CONSTRAINT IF EXISTS ai_forecasts_wbs_item_id_fkey;")
sql_lines.append("ALTER TABLE wbs_items DROP CONSTRAINT IF EXISTS wbs_items_parent_id_fkey;")
sql_lines.append("")

# Step 2: Delete existing WBS items
sql_lines.append(f"DELETE FROM wbs_items WHERE project_id = '{PROJECT_ID}';")
sql_lines.append("")

# Step 3: Insert new WBS items (deterministic UUIDs preserve allocation references)
CHUNK_SIZE = 40
for i in range(0, len(wbs_rows), CHUNK_SIZE):
    chunk = wbs_rows[i:i + CHUNK_SIZE]
    values = []
    for r in chunk:
        pid = f"'{r['parent_id']}'" if r["parent_id"] else "NULL"
        name_escaped = r["wbs_name"].replace("'", "''")
        values.append(
            f"('{r['id']}', '{PROJECT_ID}', {pid}, "
            f"'{r['wbs_code']}', '{name_escaped}', {r['qty']}, '{r['unit']}', "
            f"{r['sort_order']}, {r['level']}, {str(r['is_summary']).lower()})"
        )
    sql_lines.append(
        "INSERT INTO wbs_items (id, project_id, parent_id, wbs_code, wbs_name, qty, unit, sort_order, level, is_summary) VALUES\n"
        + ",\n".join(values) + ";"
    )
    sql_lines.append("")

# Step 4: Clean up orphaned allocations (items that no longer exist in new WBS)
sql_lines.append("-- Remove allocations referencing WBS items that no longer exist")
sql_lines.append("DELETE FROM daily_allocations WHERE wbs_item_id NOT IN (SELECT id FROM wbs_items);")
sql_lines.append("DELETE FROM baseline_snapshots WHERE wbs_item_id NOT IN (SELECT id FROM wbs_items);")
sql_lines.append("DELETE FROM ai_forecasts WHERE wbs_item_id NOT IN (SELECT id FROM wbs_items);")
sql_lines.append("")

# Step 5: Restore FK constraints
sql_lines.append("-- Restore FK constraints")
sql_lines.append("ALTER TABLE daily_allocations ADD CONSTRAINT daily_allocations_wbs_item_id_fkey FOREIGN KEY (wbs_item_id) REFERENCES wbs_items(id) ON DELETE CASCADE;")
sql_lines.append("ALTER TABLE baseline_snapshots ADD CONSTRAINT baseline_snapshots_wbs_item_id_fkey FOREIGN KEY (wbs_item_id) REFERENCES wbs_items(id) ON DELETE CASCADE;")
sql_lines.append("ALTER TABLE ai_forecasts ADD CONSTRAINT ai_forecasts_wbs_item_id_fkey FOREIGN KEY (wbs_item_id) REFERENCES wbs_items(id) ON DELETE CASCADE;")
sql_lines.append("ALTER TABLE wbs_items ADD CONSTRAINT wbs_items_parent_id_fkey FOREIGN KEY (parent_id) REFERENCES wbs_items(id) ON DELETE SET NULL;")

sql = "\n".join(sql_lines)

outfile = r"u:\Antigravity\Antigravity\Construction Schedule\wbs_rebuild_v2.sql"
with open(outfile, "w", encoding="utf-8") as f:
    f.write(sql)
print(f"Written to {outfile}")
print(f"SQL statements: {len([l for l in sql_lines if l.startswith('INSERT')])}")

# Also output as JSON for debugging
json_file = r"u:\Antigravity\Antigravity\Construction Schedule\wbs_rebuild_v2.json"
with open(json_file, "w", encoding="utf-8") as f:
    json.dump(wbs_rows, f, indent=2)
print(f"JSON written to {json_file}")
