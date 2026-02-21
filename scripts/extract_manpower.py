"""Extract daily manpower baseline from Terminplan Excel and generate SQL for daily_allocations."""

import uuid
import openpyxl
from datetime import datetime

PROJECT_ID = "5f0fc90a-00b7-4cf7-aaba-22dde8118fa1"
NAMESPACE = uuid.UUID(PROJECT_ID)

# Same UUID generation as migration script
def make_wbs_uuid(pos_code):
    return str(uuid.uuid5(NAMESPACE, pos_code))

def make_alloc_uuid(pos_code, date_str):
    return str(uuid.uuid5(NAMESPACE, f"alloc:{pos_code}:{date_str}"))

def main():
    wb = openpyxl.load_workbook(
        r"u:\Antigravity\Antigravity\Construction Schedule\DATA\Terminplan\260113_DRAFT_E2NS Terminplan.xlsx",
        data_only=True
    )
    ws = wb["E2NS Terminplan"]

    # Extract dates from Row 9, cols 27-390
    dates = {}
    for col in range(27, 391):
        v = ws.cell(row=9, column=col).value
        if isinstance(v, datetime):
            dates[col] = v.strftime("%Y-%m-%d")

    print(f"Found {len(dates)} date columns: {list(dates.values())[:5]}...{list(dates.values())[-5:]}")

    # Extract manpower data from rows 11-182
    allocations = []
    rows_with_data = 0

    for row in range(11, 183):
        pos = ws.cell(row=row, column=5).value
        if not pos:
            continue

        # Normalize pos code to match migration (e.g. "1.1.1" stays as-is)
        pos = str(pos).strip()

        wbs_item_id = make_wbs_uuid(pos)

        row_has_data = False
        for col, date_str in dates.items():
            val = ws.cell(row=row, column=col).value
            if val is not None and val != 0 and val != "":
                try:
                    manpower = float(val)
                    if manpower > 0:
                        alloc_id = make_alloc_uuid(pos, date_str)
                        allocations.append({
                            "id": alloc_id,
                            "wbs_item_id": wbs_item_id,
                            "date": date_str,
                            "planned_manpower": manpower,
                            "actual_manpower": 0,
                            "qty_done": 0,
                            "source": "baseline",
                            "pos": pos,  # for debugging
                        })
                        row_has_data = True
                except (ValueError, TypeError):
                    pass

        if row_has_data:
            rows_with_data += 1

    print(f"WBS rows with manpower data: {rows_with_data}")
    print(f"Total allocation records: {len(allocations)}")

    # Generate SQL in chunks of 40 rows (Supabase MCP limit-friendly)
    CHUNK_SIZE = 200
    chunk_idx = 0

    for i in range(0, len(allocations), CHUNK_SIZE):
        chunk = allocations[i:i + CHUNK_SIZE]
        values = []
        for a in chunk:
            values.append(
                f"('{a['id']}', '{a['wbs_item_id']}', '{a['date']}', "
                f"{a['planned_manpower']}, {a['actual_manpower']}, {a['qty_done']}, "
                f"'baseline')"
            )

        sql = (
            "INSERT INTO daily_allocations "
            "(id, wbs_item_id, date, planned_manpower, actual_manpower, qty_done, source) VALUES\n"
            + ",\n".join(values)
            + ";"
        )

        filename = f"manpower_bulk_{chunk_idx}.sql"
        with open(
            rf"u:\Antigravity\Antigravity\Construction Schedule\{filename}",
            "w", encoding="utf-8"
        ) as f:
            f.write(sql)

        print(f"  Written {filename}: {len(chunk)} rows")
        chunk_idx += 1

    print(f"\nTotal SQL files: {chunk_idx}")

    # Also write a summary of WBS codes found
    unique_pos = sorted(set(a["pos"] for a in allocations))
    print(f"\nUnique WBS codes with manpower: {len(unique_pos)}")
    for p in unique_pos[:10]:
        count = sum(1 for a in allocations if a["pos"] == p)
        print(f"  {p}: {count} days")
    print("  ...")

if __name__ == "__main__":
    main()
