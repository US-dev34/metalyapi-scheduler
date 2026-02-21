"""Insert daily manpower baseline into Supabase via Python client."""

import uuid
import openpyxl
from datetime import datetime
from supabase import create_client

SUPABASE_URL = "https://tfcmfbfnvrtsfqevwsko.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRmY21mYmZudnJ0c2ZxZXZ3c2tvIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzE1OTU3OTUsImV4cCI6MjA4NzE3MTc5NX0.Oy8IQgRFTpZI6Trl8RH15_RpofI-KqW0SddFLhvOcW0"

PROJECT_ID = "5f0fc90a-00b7-4cf7-aaba-22dde8118fa1"
NAMESPACE = uuid.UUID(PROJECT_ID)

def make_wbs_uuid(pos_code):
    return str(uuid.uuid5(NAMESPACE, pos_code))

def make_alloc_uuid(pos_code, date_str):
    return str(uuid.uuid5(NAMESPACE, f"alloc:{pos_code}:{date_str}"))

def main():
    wb = openpyxl.load_workbook(
        r"u:\Antigravity\Antigravity\Construction Schedule\DATA\Terminplan\260113_DRAFT_E2NS Terminplan.xlsx",
        data_only=True
    )
    ws = wb["E2NS Terminplan"]

    # Extract dates from Row 9, cols 27-390
    dates = {}
    for col in range(27, 391):
        v = ws.cell(row=9, column=col).value
        if isinstance(v, datetime):
            dates[col] = v.strftime("%Y-%m-%d")

    print(f"Found {len(dates)} date columns")

    # Extract manpower data
    allocations = []
    for row in range(11, 183):
        pos = ws.cell(row=row, column=5).value
        if not pos:
            continue
        pos = str(pos).strip()
        wbs_item_id = make_wbs_uuid(pos)

        for col, date_str in dates.items():
            val = ws.cell(row=row, column=col).value
            if val is not None and val != 0 and val != "":
                try:
                    manpower = float(val)
                    if manpower > 0:
                        alloc_id = make_alloc_uuid(pos, date_str)
                        allocations.append({
                            "id": alloc_id,
                            "wbs_item_id": wbs_item_id,
                            "date": date_str,
                            "planned_manpower": manpower,
                            "actual_manpower": 0,
                            "qty_done": 0,
                            "source": "baseline",
                        })
                except (ValueError, TypeError):
                    pass

    print(f"Total allocation records: {len(allocations)}")

    # Insert via Supabase client in batches of 100
    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    BATCH_SIZE = 100
    inserted = 0
    errors = 0

    for i in range(0, len(allocations), BATCH_SIZE):
        batch = allocations[i:i + BATCH_SIZE]
        try:
            result = client.table("daily_allocations").insert(batch).execute()
            inserted += len(batch)
            print(f"  Batch {i//BATCH_SIZE}: inserted {len(batch)} rows (total: {inserted})")
        except Exception as e:
            error_msg = str(e)
            # Check if it's a FK violation (wbs_item not found)
            if "violates foreign key" in error_msg:
                # Try one by one to find which ones fail
                for record in batch:
                    try:
                        client.table("daily_allocations").insert(record).execute()
                        inserted += 1
                    except Exception as e2:
                        errors += 1
                        if errors <= 5:
                            print(f"    SKIP: wbs_item_id={record['wbs_item_id']} date={record['date']} - {str(e2)[:100]}")
                print(f"  Batch {i//BATCH_SIZE}: partial insert with {errors} errors")
            else:
                print(f"  Batch {i//BATCH_SIZE} ERROR: {error_msg[:200]}")
                errors += len(batch)

    print(f"\nDone: {inserted} inserted, {errors} errors")

if __name__ == "__main__":
    main()
