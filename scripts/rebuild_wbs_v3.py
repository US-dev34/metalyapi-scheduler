"""
rebuild_wbs_v3.py — Parse WBS_E2NS_D1D2_Restructured_BACKUP_v4.xlsx
and generate SQL to rebuild wbs_items from the restructured WBS.

Reads both E2NS_WBS and D1D2_WBS sheets.
Hierarchy from WBS code depth + indent level + bold formatting.
"""

import uuid
import json
import openpyxl
from pathlib import Path

PROJECT_ID = "5f0fc90a-00b7-4cf7-aaba-22dde8118fa1"
NAMESPACE = uuid.UUID(PROJECT_ID)

EXCEL_PATH = Path(r"u:\Antigravity\Antigravity\Construction Schedule\DATA\WBS_E2NS_D1D2_Restructured_BACKUP_v4.xlsx")

def make_uuid(code: str) -> str:
    return str(uuid.uuid5(NAMESPACE, code))


def safe_num(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def get_fill_hex(cell) -> str:
    """Get fill color as hex string."""
    try:
        fill = cell.fill
        if fill and fill.start_color and fill.start_color.rgb:
            rgb = str(fill.start_color.rgb)
            if rgb != "00000000" and len(rgb) >= 6:
                return rgb
    except Exception:
        pass
    return ""


def is_bold(cell) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def get_indent(cell) -> int:
    try:
        if cell.alignment and cell.alignment.indent:
            return int(cell.alignment.indent)
    except Exception:
        pass
    return 0


def determine_level_and_summary(wbs_code: str, indent: int, bold: bool, fill_hex: str) -> tuple:
    """Determine hierarchy level and is_summary from WBS code depth + formatting."""
    parts = wbs_code.split(".")
    depth = len(parts)

    # Major sections (1.0, 1.1, 1.2, 2.1, 2.2, 2.3) = depth 2
    if depth <= 2:
        return 0 if depth == 1 else 0, True

    # Groups (1.1.1, 1.2.3, 2.1.4) = depth 3
    if depth == 3:
        # Check if this is a parent with children or a standalone leaf
        # Bold + special fill = definitely a parent
        if bold or fill_hex in ("00D6E4F0", "002E75B6"):
            return 1, True
        # Light blue fill = parent group
        if fill_hex == "00E9EFF7":
            return 1, True
        # Green fill (NTA) at depth 3 could be leaf
        if fill_hex == "00E2EFDA":
            return 1, False
        # Default depth-3 items
        return 1, False

    # Leaf items (1.1.1.1, 1.2.3.7, 2.2.4.1) = depth 4+
    return 2, False


def parse_e2ns_sheet(wb) -> list:
    """Parse E2NS_WBS sheet with 16 columns (A-P)."""
    ws = wb["E2NS_WBS"]
    rows = []

    for r in range(5, ws.max_row + 1):
        wbs_code = safe_str(ws.cell(r, 1).value)
        if not wbs_code:
            continue

        desc = safe_str(ws.cell(r, 2).value)
        if not desc:
            continue

        cell_a = ws.cell(r, 1)
        indent = get_indent(cell_a)
        bold = is_bold(cell_a) or is_bold(ws.cell(r, 2))
        fill = get_fill_hex(cell_a)

        building = safe_str(ws.cell(r, 3).value)
        nta_ref = safe_str(ws.cell(r, 4).value)
        status = safe_str(ws.cell(r, 5).value)
        budget = safe_num(ws.cell(r, 6).value)
        payment = safe_num(ws.cell(r, 7).value)
        target_kw = safe_str(ws.cell(r, 8).value)
        notes = safe_str(ws.cell(r, 9).value)
        scope = safe_str(ws.cell(r, 10).value)
        qty_ext = safe_num(ws.cell(r, 11).value)
        done_ext = safe_num(ws.cell(r, 12).value)
        rem_ext = safe_num(ws.cell(r, 13).value)
        qty_int = safe_num(ws.cell(r, 14).value)
        done_int = safe_num(ws.cell(r, 15).value)
        rem_int = safe_num(ws.cell(r, 16).value)

        level, is_summary = determine_level_and_summary(wbs_code, indent, bold, fill)

        # Total qty = ext + int (or whichever is non-zero)
        qty_total = qty_ext + qty_int

        rows.append({
            "wbs_code": wbs_code,
            "wbs_name": desc,
            "building": building,
            "nta_ref": nta_ref,
            "status": status,
            "budget_eur": budget,
            "target_kw": target_kw,
            "notes": notes,
            "scope": scope,
            "qty": qty_total,
            "qty_ext": qty_ext,
            "done_ext": done_ext,
            "rem_ext": rem_ext,
            "qty_int": qty_int,
            "done_int": done_int,
            "rem_int": rem_int,
            "unit": "pcs",
            "indent": indent,
            "bold": bold,
            "fill": fill,
            "level": level,
            "is_summary": is_summary,
            "sheet": "E2NS",
        })

    return rows


def parse_d1d2_sheet(wb) -> list:
    """Parse D1D2_WBS sheet with 9 columns (A-I)."""
    ws = wb["D1D2_WBS"]
    rows = []

    for r in range(5, ws.max_row + 1):
        wbs_code = safe_str(ws.cell(r, 1).value)
        if not wbs_code:
            continue

        desc = safe_str(ws.cell(r, 2).value)
        if not desc:
            continue

        cell_a = ws.cell(r, 1)
        indent = get_indent(cell_a)
        bold = is_bold(cell_a) or is_bold(ws.cell(r, 2))
        fill = get_fill_hex(cell_a)

        building = safe_str(ws.cell(r, 3).value)
        nta_ref = safe_str(ws.cell(r, 4).value)
        status = safe_str(ws.cell(r, 5).value)
        budget = safe_num(ws.cell(r, 6).value)
        target_kw = safe_str(ws.cell(r, 8).value)
        notes = safe_str(ws.cell(r, 9).value)

        level, is_summary = determine_level_and_summary(wbs_code, indent, bold, fill)

        rows.append({
            "wbs_code": wbs_code,
            "wbs_name": desc,
            "building": building,
            "nta_ref": nta_ref,
            "status": status,
            "budget_eur": budget,
            "target_kw": target_kw,
            "notes": notes,
            "scope": "",
            "qty": 0,
            "qty_ext": 0,
            "done_ext": 0,
            "rem_ext": 0,
            "qty_int": 0,
            "done_int": 0,
            "rem_int": 0,
            "unit": "pcs",
            "indent": indent,
            "bold": bold,
            "fill": fill,
            "level": level,
            "is_summary": is_summary,
            "sheet": "D1D2",
        })

    return rows


def determine_parent_ids(rows: list) -> list:
    """
    Assign parent_id based on WBS code hierarchy.
    1.1 is parent of 1.1.1, 1.1.1 is parent of 1.1.1.1, etc.
    """
    # Build a lookup: wbs_code -> uuid
    code_to_id = {}
    for row in rows:
        row["id"] = make_uuid(row["wbs_code"])
        code_to_id[row["wbs_code"]] = row["id"]

    # Assign parent_id by walking up the WBS code
    for row in rows:
        parts = row["wbs_code"].split(".")
        row["parent_id"] = None

        # Try progressively shorter parent codes
        for i in range(len(parts) - 1, 0, -1):
            parent_code = ".".join(parts[:i])
            if parent_code in code_to_id and parent_code != row["wbs_code"]:
                row["parent_id"] = code_to_id[parent_code]
                break

    return rows


def fix_summary_flags(rows: list):
    """Mark items as summary if they have children."""
    ids_with_children = set()
    for row in rows:
        if row["parent_id"]:
            ids_with_children.add(row["parent_id"])

    for row in rows:
        if row["id"] in ids_with_children:
            row["is_summary"] = True


def fix_levels(rows: list):
    """Set level based on WBS code depth."""
    for row in rows:
        parts = row["wbs_code"].split(".")
        depth = len(parts)
        if depth <= 2:
            row["level"] = 0
        elif depth == 3:
            row["level"] = 1
        else:
            row["level"] = 2


# ─────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────
print("=== Parsing WBS Excel v4 ===")
wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

e2ns_rows = parse_e2ns_sheet(wb)
print(f"  E2NS items: {len(e2ns_rows)}")

d1d2_rows = parse_d1d2_sheet(wb)
print(f"  D1D2 items: {len(d1d2_rows)}")

all_rows = e2ns_rows + d1d2_rows
print(f"  Total items: {len(all_rows)}")

# Assign parent IDs from WBS code hierarchy
determine_parent_ids(all_rows)
fix_summary_flags(all_rows)
fix_levels(all_rows)

# Sort by WBS code for clean ordering
all_rows.sort(key=lambda r: r["wbs_code"])

# Assign sort_order
for i, row in enumerate(all_rows):
    row["sort_order"] = (i + 1) * 100

# ─────────────────────────────────────────────────────────
# Summary
# ─────────────────────────────────────────────────────────
print(f"\n=== WBS Tree ===")
level_counts = {}
for r in all_rows:
    level_counts[r["level"]] = level_counts.get(r["level"], 0) + 1
print(f"Level breakdown: {level_counts}")

summary_count = sum(1 for r in all_rows if r["is_summary"])
leaf_count = sum(1 for r in all_rows if not r["is_summary"])
print(f"Summary rows: {summary_count}, Leaf rows: {leaf_count}")

# Print tree
for r in all_rows:
    indent = "  " * r["level"]
    summary = " [SUMMARY]" if r["is_summary"] else ""
    qty_str = f"  qty={r['qty']}" if r["qty"] > 0 else ""
    scope_str = f"  scope={r['scope']}" if r["scope"] else ""
    bldg_str = f"  [{r['building']}]" if r["building"] else ""
    print(f"  {indent}{r['wbs_code']:14s} {r['wbs_name'][:50]}{summary}{qty_str}{scope_str}{bldg_str}")

# ─────────────────────────────────────────────────────────
# Generate SQL
# ─────────────────────────────────────────────────────────
print("\n=== Generating SQL ===")

sql_lines = []

# Step 1: Drop FK constraints to prevent CASCADE deletes
sql_lines.append("-- Temporarily drop FK constraints to prevent cascade deletion of allocations")
sql_lines.append("ALTER TABLE daily_allocations DROP CONSTRAINT IF EXISTS daily_allocations_wbs_item_id_fkey;")
sql_lines.append("ALTER TABLE baseline_snapshots DROP CONSTRAINT IF EXISTS baseline_snapshots_wbs_item_id_fkey;")
sql_lines.append("ALTER TABLE ai_forecasts DROP CONSTRAINT IF EXISTS ai_forecasts_wbs_item_id_fkey;")
sql_lines.append("ALTER TABLE wbs_items DROP CONSTRAINT IF EXISTS wbs_items_parent_id_fkey;")
sql_lines.append("")

# Step 2: Delete existing WBS items
sql_lines.append(f"DELETE FROM wbs_items WHERE project_id = '{PROJECT_ID}';")
sql_lines.append("")

# Step 3: Insert new WBS items
CHUNK_SIZE = 30
insert_cols = (
    "id, project_id, parent_id, wbs_code, wbs_name, qty, unit, sort_order, level, is_summary, "
    "building, nta_ref, status, scope, budget_eur, target_kw, notes, "
    "qty_ext, done_ext, rem_ext, qty_int, done_int, rem_int"
)

for i in range(0, len(all_rows), CHUNK_SIZE):
    chunk = all_rows[i:i + CHUNK_SIZE]
    values = []
    for r in chunk:
        pid = f"'{r['parent_id']}'" if r["parent_id"] else "NULL"
        name_esc = r["wbs_name"].replace("'", "''")
        notes_esc = r["notes"].replace("'", "''") if r["notes"] else ""
        building = f"'{r['building']}'" if r["building"] else "NULL"
        nta_ref = f"'{r['nta_ref']}'" if r["nta_ref"] else "NULL"
        status = f"'{r['status']}'" if r["status"] else "NULL"
        scope = f"'{r['scope']}'" if r["scope"] else "NULL"
        target_kw = f"'{r['target_kw']}'" if r["target_kw"] else "NULL"
        notes_val = f"'{notes_esc}'" if notes_esc else "NULL"

        values.append(
            f"('{r['id']}', '{PROJECT_ID}', {pid}, '{r['wbs_code']}', '{name_esc}', "
            f"{r['qty']}, '{r['unit']}', {r['sort_order']}, {r['level']}, {str(r['is_summary']).lower()}, "
            f"{building}, {nta_ref}, {status}, {scope}, {r['budget_eur']}, {target_kw}, {notes_val}, "
            f"{r['qty_ext']}, {r['done_ext']}, {r['rem_ext']}, {r['qty_int']}, {r['done_int']}, {r['rem_int']})"
        )

    sql_lines.append(
        f"INSERT INTO wbs_items ({insert_cols}) VALUES\n"
        + ",\n".join(values) + ";"
    )
    sql_lines.append("")

# Step 4: Clean up orphaned allocations
sql_lines.append("-- Remove allocations referencing WBS items that no longer exist")
sql_lines.append("DELETE FROM daily_allocations WHERE wbs_item_id NOT IN (SELECT id FROM wbs_items);")
sql_lines.append("DELETE FROM baseline_snapshots WHERE wbs_item_id NOT IN (SELECT id FROM wbs_items);")
sql_lines.append("DELETE FROM ai_forecasts WHERE wbs_item_id NOT IN (SELECT id FROM wbs_items);")
sql_lines.append("")

# Step 5: Restore FK constraints
sql_lines.append("-- Restore FK constraints")
sql_lines.append("ALTER TABLE daily_allocations ADD CONSTRAINT daily_allocations_wbs_item_id_fkey FOREIGN KEY (wbs_item_id) REFERENCES wbs_items(id) ON DELETE CASCADE;")
sql_lines.append("ALTER TABLE baseline_snapshots ADD CONSTRAINT baseline_snapshots_wbs_item_id_fkey FOREIGN KEY (wbs_item_id) REFERENCES wbs_items(id) ON DELETE CASCADE;")
sql_lines.append("ALTER TABLE ai_forecasts ADD CONSTRAINT ai_forecasts_wbs_item_id_fkey FOREIGN KEY (wbs_item_id) REFERENCES wbs_items(id) ON DELETE CASCADE;")
sql_lines.append("ALTER TABLE wbs_items ADD CONSTRAINT wbs_items_parent_id_fkey FOREIGN KEY (parent_id) REFERENCES wbs_items(id) ON DELETE SET NULL;")

sql = "\n".join(sql_lines)

outfile = Path(r"u:\Antigravity\Antigravity\Construction Schedule\wbs_rebuild_v3.sql")
with open(outfile, "w", encoding="utf-8") as f:
    f.write(sql)
print(f"\nWritten to {outfile}")
print(f"SQL INSERT statements: {len([l for l in sql_lines if l.startswith('INSERT')])}")

# Also output JSON
json_file = Path(r"u:\Antigravity\Antigravity\Construction Schedule\wbs_rebuild_v3.json")
with open(json_file, "w", encoding="utf-8") as f:
    json.dump(all_rows, f, indent=2, default=str)
print(f"JSON written to {json_file}")
